"""Intentionally naive full-workbook MCP baseline for honest comparison."""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

from mcp.server.fastmcp import FastMCP
from mcp.types import ToolAnnotations
from openpyxl import load_workbook

server = FastMCP(
    "Excel naive dump baseline",
    instructions="Read workbook or sheet CSV dumps. This server has no semantic index.",
    json_response=True,
)
READ_ANNOTATIONS = ToolAnnotations(readOnlyHint=True, openWorldHint=False)


def _render(value: Any, cached: Any) -> Any:
    if isinstance(value, str) and value.startswith("="):
        return f"{value} => {cached!s}" if cached is not None else value
    return value


def _sheet_csv(path: Path, sheet: str) -> str:
    formulas = load_workbook(path, read_only=True, data_only=False, keep_links=True)
    values = load_workbook(path, read_only=True, data_only=True, keep_links=True)
    try:
        if sheet not in formulas.sheetnames:
            raise ValueError(f"unknown worksheet: {sheet}")
        formula_sheet = formulas[sheet]
        value_sheet = values[sheet]
        output = io.StringIO(newline="")
        writer = csv.writer(output, lineterminator="\n")
        for formula_row, value_row in zip(
            formula_sheet.iter_rows(values_only=True),
            value_sheet.iter_rows(values_only=True),
            strict=True,
        ):
            writer.writerow(
                _render(formula, cached)
                for formula, cached in zip(formula_row, value_row, strict=True)
            )
        return output.getvalue()
    finally:
        formulas.close()
        values.close()


def read_sheet(path: str, sheet: str) -> dict[str, str]:
    """Return one complete worksheet as CSV text, including formula caches."""
    workbook = Path(path).expanduser().resolve(strict=True)
    return {"sheet": sheet, "csv": _sheet_csv(workbook, sheet)}


def read_workbook_full(path: str) -> dict[str, Any]:
    """Return every worksheet as CSV text with no semantic reduction."""
    workbook = Path(path).expanduser().resolve(strict=True)
    probe = load_workbook(workbook, read_only=True, data_only=False, keep_links=True)
    try:
        names = tuple(probe.sheetnames)
    finally:
        probe.close()
    return {
        "workbook": workbook.name,
        "sheets": [{"sheet": name, "csv": _sheet_csv(workbook, name)} for name in names],
    }


server.tool(annotations=READ_ANNOTATIONS)(read_workbook_full)
server.tool(annotations=READ_ANNOTATIONS)(read_sheet)


def main() -> None:
    server.run(transport="stdio")


if __name__ == "__main__":
    main()
