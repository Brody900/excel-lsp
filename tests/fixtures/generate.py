"""Build the deterministic Excel workbooks used by the test suite.

openpyxl is intentionally limited to authoring new fixture workbooks. The
post-processing helpers below patch raw OOXML so formula caches and shared
formula groups match files saved by Excel, then repack every archive with a
stable member order and timestamp.
"""

from __future__ import annotations

import math
import posixpath
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Literal
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import lxml.etree as etree
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

GENERATED_DIR = Path(__file__).resolve().parent / "generated"

_DOCUMENT_TIMESTAMP = datetime(2000, 1, 1, 0, 0, 0)
_DOCUMENT_TIMESTAMP_TEXT = "2000-01-01T00:00:00Z"
_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
_DCTERMS_NS = "http://purl.org/dc/terms/"
_REL_TYPE_BASE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_VBA_PROJECT = Path(__file__).resolve().parent / "assets" / "vbaProject.bin"
_PNG_16X16 = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000100000001008060000001ff3ff61"
    "0000001d49444154789c63601805a360148c02c40f181818f8cfc0c000000c8c01"
    "1f90f52f5a0000000049454e44ae426082"
)

CachedValueKind = Literal["number", "string", "error", "boolean"]


@dataclass(frozen=True, slots=True)
class CachedValue:
    """A formula result known by the fixture author, not a general evaluator."""

    value: int | float | str | bool
    kind: CachedValueKind = "number"


@dataclass(frozen=True, slots=True)
class SharedFormulaGroup:
    """One OOXML shared-formula rectangle and its top-left master cell."""

    range_ref: str
    master_ref: str
    shared_index: int


def _read_archive(path: Path) -> dict[str, bytes]:
    with ZipFile(path, "r") as archive:
        return {name: archive.read(name) for name in archive.namelist()}


def _write_deterministic_archive(path: Path, members: Mapping[str, bytes]) -> None:
    stable_members = dict(members)
    core_properties = stable_members.get("docProps/core.xml")
    if core_properties is not None:
        core = etree.fromstring(core_properties)
        for property_name in ("created", "modified"):
            element = core.find(f"{{{_DCTERMS_NS}}}{property_name}")
            if element is not None:
                element.text = _DOCUMENT_TIMESTAMP_TEXT
        stable_members["docProps/core.xml"] = etree.tostring(core, encoding="utf-8")

    temporary = path.with_name(f".{path.name}.tmp")
    temporary.unlink(missing_ok=True)
    try:
        with ZipFile(temporary, "w", compression=ZIP_DEFLATED, compresslevel=9) as archive:
            for name in sorted(stable_members):
                info = ZipInfo(filename=name, date_time=_ZIP_TIMESTAMP)
                info.compress_type = ZIP_DEFLATED
                info.create_system = 0
                info.external_attr = 0
                info.internal_attr = 0
                info.extra = b""
                info.comment = b""
                archive.writestr(
                    info,
                    stable_members[name],
                    compress_type=ZIP_DEFLATED,
                    compresslevel=9,
                )
        temporary.replace(path)
    finally:
        temporary.unlink(missing_ok=True)


def repack_deterministic(path: Path) -> None:
    """Sort ZIP members and apply the earliest valid ZIP timestamp to each."""
    _write_deterministic_archive(path, _read_archive(path))


def _worksheet_part(members: Mapping[str, bytes], sheet_name: str) -> str:
    workbook = etree.fromstring(members["xl/workbook.xml"])
    sheet = next(
        (
            candidate
            for candidate in workbook.findall(f".//{{{_MAIN_NS}}}sheet")
            if candidate.get("name") == sheet_name
        ),
        None,
    )
    if sheet is None:
        raise ValueError(f"worksheet not found: {sheet_name}")

    relationship_id = sheet.get(f"{{{_OFFICE_REL_NS}}}id")
    relationships = etree.fromstring(members["xl/_rels/workbook.xml.rels"])
    relationship = next(
        (
            candidate
            for candidate in relationships.findall(f"{{{_PACKAGE_REL_NS}}}Relationship")
            if candidate.get("Id") == relationship_id
        ),
        None,
    )
    if relationship is None or not relationship.get("Target"):
        raise ValueError(f"worksheet relationship is missing: {sheet_name}")

    target = relationship.get("Target")
    assert target is not None
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join("xl", target))


def _serialized_cached_value(cached: CachedValue) -> tuple[str | None, str]:
    if cached.kind == "number":
        if isinstance(cached.value, bool) or not isinstance(cached.value, (int, float)):
            raise TypeError("numeric cached values must be int or float, not bool")
        if isinstance(cached.value, float) and not math.isfinite(cached.value):
            raise ValueError("numeric cached values must be finite")
        return None, str(cached.value)
    if cached.kind == "boolean":
        if not isinstance(cached.value, bool):
            raise TypeError("boolean cached values must be bool")
        return "b", "1" if cached.value else "0"
    if not isinstance(cached.value, str):
        raise TypeError(f"{cached.kind} cached values must be strings")
    return ("str" if cached.kind == "string" else "e"), cached.value


def inject_cached_values(
    workbook_path: Path,
    sheet_name: str,
    cached_values: Mapping[str, CachedValue],
) -> None:
    """Inject fixture-author-computed formula results into one worksheet part."""
    members = _read_archive(workbook_path)
    part = _worksheet_part(members, sheet_name)
    worksheet = etree.fromstring(members[part])
    cells = {
        cell.get("r"): cell for cell in worksheet.findall(f".//{{{_MAIN_NS}}}c") if cell.get("r")
    }

    for ref, cached in cached_values.items():
        cell = cells.get(ref)
        if cell is None:
            raise ValueError(f"cannot cache missing cell: {sheet_name}!{ref}")
        formula = cell.find(f"{{{_MAIN_NS}}}f")
        if formula is None:
            raise ValueError(f"cannot cache non-formula cell: {sheet_name}!{ref}")

        cell_type, serialized = _serialized_cached_value(cached)
        if cell_type is None:
            cell.attrib.pop("t", None)
        else:
            cell.set("t", cell_type)

        for existing in cell.findall(f"{{{_MAIN_NS}}}v"):
            cell.remove(existing)
        value = etree.Element(f"{{{_MAIN_NS}}}v")
        value.text = serialized
        cell.insert(cell.index(formula) + 1, value)

    members[part] = etree.tostring(worksheet, encoding="utf-8")
    _write_deterministic_archive(workbook_path, members)


def _refs_in_range(range_ref: str) -> tuple[str, ...]:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ValueError(f"shared formula range must be bounded: {range_ref}")
    return tuple(
        f"{get_column_letter(column)}{row}"
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    )


def convert_shared_formula_groups(
    workbook_path: Path,
    sheet_name: str,
    groups: Sequence[SharedFormulaGroup],
) -> None:
    """Convert full formulas to OOXML shared masters and si-only followers."""
    members = _read_archive(workbook_path)
    part = _worksheet_part(members, sheet_name)
    worksheet = etree.fromstring(members[part])
    cells = {
        cell.get("r"): cell for cell in worksheet.findall(f".//{{{_MAIN_NS}}}c") if cell.get("r")
    }

    seen_indexes: set[int] = set()
    for group in groups:
        if group.shared_index in seen_indexes:
            raise ValueError(f"duplicate shared formula index: {group.shared_index}")
        seen_indexes.add(group.shared_index)

        refs = _refs_in_range(group.range_ref)
        if not refs or refs[0] != group.master_ref:
            raise ValueError("shared formula master must be the top-left cell of its range")

        formulas = {}
        for ref in refs:
            cell = cells.get(ref)
            formula = None if cell is None else cell.find(f"{{{_MAIN_NS}}}f")
            if formula is None or not formula.text:
                raise ValueError(f"shared formula source is missing: {sheet_name}!{ref}")
            formulas[ref] = formula

        master = formulas[group.master_ref]
        master.attrib.clear()
        master.set("t", "shared")
        master.set("ref", group.range_ref)
        master.set("si", str(group.shared_index))

        for ref in refs[1:]:
            follower = formulas[ref]
            follower.attrib.clear()
            follower.set("t", "shared")
            follower.set("si", str(group.shared_index))
            follower.text = None

    members[part] = etree.tostring(worksheet, encoding="utf-8")
    _write_deterministic_archive(workbook_path, members)


def inject_external_link(
    workbook_path: Path,
    *,
    target: str,
    relationship_id: str = "rIdExternalLink1",
) -> None:
    """Add one genuine [1] external-link map and relationship target."""
    members = _read_archive(workbook_path)
    workbook = etree.fromstring(members["xl/workbook.xml"])
    sheets = workbook.find(f"{{{_MAIN_NS}}}sheets")
    if sheets is None:
        raise ValueError("workbook is missing its sheet catalog")
    external_references = etree.Element(f"{{{_MAIN_NS}}}externalReferences")
    external_reference = etree.SubElement(
        external_references,
        f"{{{_MAIN_NS}}}externalReference",
    )
    external_reference.set(f"{{{_OFFICE_REL_NS}}}id", relationship_id)
    workbook.insert(workbook.index(sheets) + 1, external_references)

    relationships = etree.fromstring(members["xl/_rels/workbook.xml.rels"])
    etree.SubElement(
        relationships,
        f"{{{_PACKAGE_REL_NS}}}Relationship",
        Id=relationship_id,
        Type=f"{_REL_TYPE_BASE}/externalLink",
        Target="externalLinks/externalLink1.xml",
    )

    external_link = etree.Element(
        f"{{{_MAIN_NS}}}externalLink",
        nsmap={"r": _OFFICE_REL_NS},
    )
    external_book = etree.SubElement(external_link, f"{{{_MAIN_NS}}}externalBook")
    external_book.set(f"{{{_OFFICE_REL_NS}}}id", "rIdExternalBook1")
    external_relationships = etree.Element(f"{{{_PACKAGE_REL_NS}}}Relationships")
    etree.SubElement(
        external_relationships,
        f"{{{_PACKAGE_REL_NS}}}Relationship",
        Id="rIdExternalBook1",
        Type=f"{_REL_TYPE_BASE}/externalLinkPath",
        Target=target,
        TargetMode="External",
    )

    content_types = etree.fromstring(members["[Content_Types].xml"])
    etree.SubElement(
        content_types,
        f"{{{_CONTENT_TYPES_NS}}}Override",
        PartName="/xl/externalLinks/externalLink1.xml",
        ContentType=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"
        ),
    )

    members["xl/workbook.xml"] = etree.tostring(workbook, encoding="utf-8")
    members["xl/_rels/workbook.xml.rels"] = etree.tostring(
        relationships,
        encoding="utf-8",
    )
    members["xl/externalLinks/externalLink1.xml"] = etree.tostring(
        external_link,
        encoding="utf-8",
    )
    members["xl/externalLinks/_rels/externalLink1.xml.rels"] = etree.tostring(
        external_relationships,
        encoding="utf-8",
    )
    members["[Content_Types].xml"] = etree.tostring(content_types, encoding="utf-8")
    _write_deterministic_archive(workbook_path, members)


def inject_vba_project(workbook_path: Path, project_path: Path = _VBA_PROJECT) -> None:
    """Inject the sanctioned F16 VBA project and macro-enabled package metadata."""
    members = _read_archive(workbook_path)
    relationships = etree.fromstring(members["xl/_rels/workbook.xml.rels"])
    if any(
        relationship.get("Id") == "rIdVbaProject"
        for relationship in relationships.findall(f"{{{_PACKAGE_REL_NS}}}Relationship")
    ):
        raise ValueError("workbook already contains the fixture VBA relationship id")
    etree.SubElement(
        relationships,
        f"{{{_PACKAGE_REL_NS}}}Relationship",
        Id="rIdVbaProject",
        Type=f"{_REL_TYPE_BASE}/vbaProject",
        Target="vbaProject.bin",
    )

    content_types = etree.fromstring(members["[Content_Types].xml"])
    workbook_override = next(
        (
            override
            for override in content_types.findall(f"{{{_CONTENT_TYPES_NS}}}Override")
            if override.get("PartName") == "/xl/workbook.xml"
        ),
        None,
    )
    if workbook_override is None:
        raise ValueError("workbook content-type override is missing")
    workbook_override.set(
        "ContentType",
        "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    )
    etree.SubElement(
        content_types,
        f"{{{_CONTENT_TYPES_NS}}}Override",
        PartName="/xl/vbaProject.bin",
        ContentType="application/vnd.ms-office.vbaProject",
    )

    members["xl/_rels/workbook.xml.rels"] = etree.tostring(relationships)
    members["[Content_Types].xml"] = etree.tostring(content_types)
    members["xl/vbaProject.bin"] = project_path.read_bytes()
    _write_deterministic_archive(workbook_path, members)


def _new_workbook(sheet_name: str) -> tuple[Workbook, Worksheet]:
    workbook = Workbook()
    worksheet = workbook.active
    if not isinstance(worksheet, Worksheet):
        raise RuntimeError("new workbook did not create a worksheet")
    worksheet.title = sheet_name
    workbook.properties.creator = "Excel LSP fixture generator"
    workbook.properties.lastModifiedBy = "Excel LSP fixture generator"
    workbook.properties.created = _DOCUMENT_TIMESTAMP
    workbook.properties.modified = _DOCUMENT_TIMESTAMP
    workbook.properties.title = f"Excel LSP {sheet_name} fixture"
    return workbook, worksheet


def _table(name: str, ref: str) -> Table:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    return table


def _style_header_cells(
    worksheet: Worksheet,
    refs: Sequence[str],
    *,
    fill_color: str = "1F4E78",
    font_color: str = "FFFFFF",
) -> None:
    """Apply an explicit format shift for deterministic header inference."""
    font = Font(bold=True, color=font_color)
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    for ref in refs:
        worksheet[ref].font = font
        worksheet[ref].fill = fill


def _generate_f01(output_dir: Path) -> Path:
    path = output_dir / "basic_single_table.xlsx"
    workbook, worksheet = _new_workbook("Sales")
    worksheet.append(("Item", "Quantity", "UnitPrice", "LineTotal"))

    rows = (
        ("Widget", 2, 3.5),
        ("Gadget", 5, 1.25),
        ("Bracket", 3, 4.0),
        ("Cable", 7, 2.0),
        ("Adapter", 4, 6.75),
    )
    caches: dict[str, CachedValue] = {}
    for row_number, (item, quantity, unit_price) in enumerate(rows, start=2):
        worksheet.append((item, quantity, unit_price, f"=B{row_number}*C{row_number}"))
        caches[f"D{row_number}"] = CachedValue(quantity * unit_price)

    worksheet.add_table(_table("SalesTable", "A1:D6"))
    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "Sales", caches)
    repack_deterministic(path)
    return path


def _generate_f02(output_dir: Path) -> Path:
    path = output_dir / "multi_region.xlsx"
    workbook, worksheet = _new_workbook("Islands")

    first_island = {
        1: ("Product", "Units", "UnitPrice"),
        2: ("Widget", 4, 2.5),
        3: ("Gadget", 3, 4.0),
        5: ("Cable", 8, 1.25),
        6: ("Adapter", 2, 6.75),
    }
    for row_number, values in first_island.items():
        for column_number, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column_number, value=value)

    second_island = (
        ("Department", "Budget", "Active"),
        ("Sales", 120000.5, True),
        ("Support", 85000.25, True),
        ("Research", 99000.75, False),
        ("Operations", 101500.0, True),
    )
    for row_number, values in enumerate(second_island, start=2):
        for column_number, value in enumerate(values, start=6):
            worksheet.cell(row=row_number, column=column_number, value=value)

    third_island = (
        ("Metric", "Value", "Flag"),
        ("LatencyMs", 12.5, "ok"),
        ("Errors", 2, "review"),
        ("Throughput", 450, "ok"),
    )
    for row_number, values in enumerate(third_island, start=10):
        for column_number, value in enumerate(values, start=2):
            worksheet.cell(row=row_number, column=column_number, value=value)

    _style_header_cells(worksheet, ("A1", "B1", "C1"))
    _style_header_cells(worksheet, ("F2", "G2", "H2"))
    _style_header_cells(worksheet, ("B10", "C10", "D10"))
    workbook.save(path)
    workbook.close()
    repack_deterministic(path)
    return path


def _generate_f03(output_dir: Path) -> Path:
    path = output_dir / "cross_sheet_model.xlsx"
    workbook, inputs = _new_workbook("Inputs")
    calc = workbook.create_sheet("Calc")
    summary = workbook.create_sheet("Summary")

    inputs.append(("Input", "Value", "Unit"))
    inputs.append(("GrowthRate", 0.10, "%"))
    inputs.append(("BaseRevenue", 1000, "USD"))
    inputs.append(("CostRate", 0.60, "%"))
    inputs.append(("TaxRate", 0.25, "%"))
    inputs.add_table(_table("InputsTable", "A1:C5"))

    revenues: tuple[int | float, ...] = (1000, 1100, 1210, 1331, 1464.1)
    costs: tuple[int | float, ...] = (600, 660, 726, 798.6, 878.46)
    after_tax: tuple[int | float, ...] = (300, 330, 363, 399.3, 439.23)
    calc.append(("Year", "Revenue", "Cost", "AfterTax"))
    calc_caches: dict[str, CachedValue] = {}
    for offset, (revenue, cost, result) in enumerate(
        zip(revenues, costs, after_tax, strict=True),
        start=2,
    ):
        revenue_formula = "=Inputs!$B$3" if offset == 2 else f"=B{offset - 1}*(1+Inputs!$B$2)"
        calc.append(
            (
                2024 + offset,
                revenue_formula,
                f"=B{offset}*Inputs!$B$4",
                f"=(B{offset}-C{offset})*(1-Inputs!$B$5)",
            )
        )
        calc_caches[f"B{offset}"] = CachedValue(revenue)
        calc_caches[f"C{offset}"] = CachedValue(cost)
        calc_caches[f"D{offset}"] = CachedValue(result)
    calc.add_table(_table("CalcTable", "A1:D6"))

    summary.append(("Section", "Metric", "Value"))
    summary_caches: dict[str, CachedValue] = {}
    for row_number, result in enumerate(after_tax, start=2):
        year = 2024 + row_number
        summary.append(("Annual", f"FY{year}", f"=Calc!D{row_number}"))
        summary_caches[f"C{row_number}"] = CachedValue(result)
    summary.append(("KPI", "Starting Revenue", "=Calc!B2"))
    summary.append(("KPI", "Ending Revenue", "=Calc!B6"))
    summary.append(("KPI", "Total Cost", "=SUM(Calc!C2:C6)"))
    summary.append(("KPI", "Total After Tax", "=SUM(Calc!D2:D6)"))
    summary_caches.update(
        {
            "C7": CachedValue(1000),
            "C8": CachedValue(1464.1),
            "C9": CachedValue(3663.06),
            "C10": CachedValue(1831.53),
        }
    )
    summary.add_table(_table("SummaryTable", "A1:C10"))

    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "Calc", calc_caches)
    inject_cached_values(path, "Summary", summary_caches)
    repack_deterministic(path)
    return path


def _generate_f04(output_dir: Path) -> Path:
    path = output_dir / "named_ranges.xlsx"
    workbook, inputs = _new_workbook("Inputs")
    calc = workbook.create_sheet("Calc")

    inputs.append(("Input", "Value"))
    inputs.append(("BaseAmount", 100))
    inputs.append(("GlobalRate", 0.10))
    inputs.append(("Global projection", "=BaseAmount*(1+GlobalRate)"))

    calc.append(("Input", "Value"))
    calc.append(("ScopedRate", 0.05))
    calc.append(("Base amount", "=BaseAmount"))
    calc.append(("Scoped projection", "=BaseAmount*(1+ScopedRate)"))

    workbook.defined_names.add(DefinedName("BaseAmount", attr_text="'Inputs'!$B$2"))
    workbook.defined_names.add(DefinedName("GlobalRate", attr_text="'Inputs'!$B$3"))
    workbook.defined_names.add(DefinedName("ScopedRate", attr_text="'Calc'!$B$2", localSheetId=1))

    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "Inputs", {"B4": CachedValue(110)})
    inject_cached_values(
        path,
        "Calc",
        {
            "B3": CachedValue(100),
            "B4": CachedValue(105),
        },
    )
    repack_deterministic(path)
    return path


def _generate_f05(output_dir: Path) -> Path:
    path = output_dir / "structured_table.xlsx"
    workbook, worksheet = _new_workbook("Structured")
    worksheet.append(("Product", "Qty", "Price", "LineTotal", None, "GrandTotal"))

    rows = (
        ("Widget", 2, 3.5),
        ("Gadget", 5, 1.25),
        ("Cable", 7, 2.0),
        ("Adapter", 4, 6.75),
    )
    caches: dict[str, CachedValue] = {}
    for row_number, (product, quantity, price) in enumerate(rows, start=2):
        worksheet.cell(row=row_number, column=1, value=product)
        worksheet.cell(row=row_number, column=2, value=quantity)
        worksheet.cell(row=row_number, column=3, value=price)
        worksheet.cell(row=row_number, column=4, value="=[@Qty]*[@Price]")
        caches[f"D{row_number}"] = CachedValue(quantity * price)

    worksheet["A6"] = "Totals"
    worksheet["B6"] = "=SUBTOTAL(109,Table1[Qty])"
    worksheet["D6"] = "=SUBTOTAL(109,Table1[LineTotal])"
    worksheet["F2"] = "=SUM(Table1[LineTotal])"
    caches.update(
        {
            "B6": CachedValue(18),
            "D6": CachedValue(54.25),
            "F2": CachedValue(54.25),
        }
    )

    table = _table("Table1", "A1:D6")
    table.totalsRowCount = 1
    table.autoFilter = AutoFilter(ref=table.ref)
    table.tableColumns = [
        TableColumn(id=column_id, name=name)
        for column_id, name in enumerate(
            ("Product", "Qty", "Price", "LineTotal"),
            start=1,
        )
    ]
    table.tableColumns[0].totalsRowLabel = "Totals"
    table.tableColumns[1].totalsRowFunction = "sum"
    table.tableColumns[3].totalsRowFunction = "sum"
    worksheet.add_table(table)

    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "Structured", caches)
    repack_deterministic(path)
    return path


def _generate_f07(output_dir: Path) -> Path:
    path = output_dir / "formula_blocks.xlsx"
    workbook, worksheet = _new_workbook("FormulaBlocks")
    worksheet.append(("Factor", "Multiplier", "Product"))

    caches: dict[str, CachedValue] = {}
    for row_number in range(2, 22):
        factor = row_number - 1
        multiplier = (row_number % 5) + 2
        formula = (
            f"=A{row_number}+B{row_number}" if row_number == 12 else f"=A{row_number}*B{row_number}"
        )
        worksheet.append((factor, multiplier, formula))
        expected = factor + multiplier if row_number == 12 else factor * multiplier
        caches[f"C{row_number}"] = CachedValue(expected)

    worksheet.add_table(_table("FormulaBlocksTable", "A1:C21"))
    worksheet["E1"] = "Read-only metadata probe"
    worksheet.merge_cells("E1:F1")
    workbook.save(path)
    workbook.close()

    inject_cached_values(path, "FormulaBlocks", caches)
    convert_shared_formula_groups(
        path,
        "FormulaBlocks",
        (
            SharedFormulaGroup("C2:C11", "C2", 0),
            SharedFormulaGroup("C13:C21", "C13", 1),
        ),
    )
    repack_deterministic(path)
    return path


def _generate_f08(output_dir: Path) -> Path:
    path = output_dir / "errors.xlsx"
    workbook, worksheet = _new_workbook("Errors")
    worksheet.append(("ExpectedError", "FormulaResult"))
    errors = (
        "#REF!",
        "#DIV/0!",
        "#N/A",
        "#VALUE!",
        "#NAME?",
        "#NUM!",
        "#SPILL!",
        "#CALC!",
        "#BLOCKED!",
        "#FIELD!",
    )
    cached_values: dict[str, CachedValue] = {}
    for row_number, error_value in enumerate(errors, start=2):
        worksheet.append((f"Expected {error_value}", "=NA()"))
        cached_values[f"B{row_number}"] = CachedValue(error_value, "error")
    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "Errors", cached_values)
    repack_deterministic(path)
    return path


def _generate_f10(output_dir: Path) -> Path:
    path = output_dir / "external_link.xlsx"
    workbook, worksheet = _new_workbook("External")
    worksheet.append(("LinkedValue",))
    worksheet["A2"] = "=[1]Data!A1"
    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "External", {"A2": CachedValue(0)})
    inject_external_link(path, target="missing/linked-budget.xlsx")
    repack_deterministic(path)
    return path


def _generate_f11(output_dir: Path) -> Path:
    path = output_dir / "indirect_offset.xlsx"
    workbook, worksheet = _new_workbook("DynamicRefs")
    worksheet.append(("Input", "Indirect", "Offset"))
    worksheet.append((10, '=INDIRECT("A2")', "=OFFSET(A2,1,0)"))
    worksheet.append((20, None, None))
    workbook.save(path)
    workbook.close()
    inject_cached_values(
        path,
        "DynamicRefs",
        {
            "B2": CachedValue(10),
            "C2": CachedValue(20),
        },
    )
    repack_deterministic(path)
    return path


def _generate_f18(output_dir: Path) -> Path:
    path = output_dir / "volatile.xlsx"
    workbook, worksheet = _new_workbook("Volatile")
    worksheet.append(("Function", "Value"))
    worksheet.append(("NOW", "=NOW()"))
    worksheet.append(("RAND", "=RAND()"))
    workbook.save(path)
    workbook.close()
    inject_cached_values(
        path,
        "Volatile",
        {
            "B2": CachedValue(45292.5),
            "B3": CachedValue(0.25),
        },
    )
    repack_deterministic(path)
    return path


def _generate_f09a(output_dir: Path) -> Path:
    path = output_dir / "circular.xlsx"
    workbook, worksheet = _new_workbook("Circular")
    worksheet.append(("Node", "Value"))
    worksheet.append(("A", "=B3+1"))
    worksheet.append(("B", "=B2+1"))
    workbook.save(path)
    workbook.close()
    inject_cached_values(
        path,
        "Circular",
        {
            "B2": CachedValue(0),
            "B3": CachedValue(0),
        },
    )
    repack_deterministic(path)
    return path


def _generate_f09b(output_dir: Path) -> Path:
    path = output_dir / "running_total.xlsx"
    workbook, worksheet = _new_workbook("RunningTotal")
    worksheet.append(("Row", "RunningTotal"))
    worksheet.append((0, 0))

    caches: dict[str, CachedValue] = {}
    for row_number in range(3, 50_003):
        worksheet.cell(row=row_number, column=1, value=row_number - 2)
        worksheet.cell(
            row=row_number,
            column=2,
            value=f"=SUM($B$2:B{row_number - 1})",
        )
        caches[f"B{row_number}"] = CachedValue(0)

    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "RunningTotal", caches)
    repack_deterministic(path)
    return path


def _generate_f12(output_dir: Path) -> Path:
    path = output_dir / "merged_headers.xlsx"
    workbook, worksheet = _new_workbook("MergedHeaders")

    worksheet["A1"] = "Region"
    worksheet["B1"] = "Revenue"
    worksheet["E1"] = "Units"
    worksheet["B2"] = "Q1"
    worksheet["C2"] = "Q2"
    worksheet["D2"] = "Q3"
    worksheet["E2"] = "Actual"
    worksheet["F2"] = "Target"
    worksheet.merge_cells("A1:A2")
    worksheet.merge_cells("B1:D1")
    worksheet.merge_cells("E1:F1")

    rows = (
        ("North", 100.5, 120.25, 130.75, 10, 12),
        ("South", 90.25, 98.5, 105.75, 9, 10),
        ("West", 110.75, 115.5, 125.25, 11, 13),
        ("East", 80.5, 88.75, 95.25, 8, 9),
    )
    for row_number, values in enumerate(rows, start=3):
        for column_number, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column_number, value=value)

    _style_header_cells(worksheet, ("A1", "B1", "E1"))
    _style_header_cells(
        worksheet,
        ("B2", "C2", "D2", "E2", "F2"),
        fill_color="D9EAF7",
        font_color="000000",
    )
    workbook.save(path)
    workbook.close()
    repack_deterministic(path)
    return path


def _generate_f13(output_dir: Path) -> Path:
    path = output_dir / "mixed_types.xlsx"
    workbook, worksheet = _new_workbook("MixedTypes")
    worksheet.append(
        (
            "RecordID",
            "PostingDate",
            "Amount",
            "MarginPct",
            "AccountCode",
            "Approved",
            "MixedSample",
        )
    )
    rows = (
        (1, 45292, 1250.50, 0.125, "00100", True, 10),
        (2, 45323, -42.75, 0.05, "00101", False, "pending"),
        (3, 45352, 0.50, 0.20, "00102", True, 30),
        (4, 45383, 999.99, 0.075, "00103", False, "hold"),
        (5, 45413, 10.25, 0.0, "00104", True, 50),
        (6, 45444, 200.10, 1.0, "00105", False, "done"),
    )
    for values in rows:
        worksheet.append(values)

    for row_number in range(2, 8):
        worksheet[f"B{row_number}"].number_format = "yyyy-mm-dd" if row_number <= 4 else "mm-dd-yy"
        worksheet[f"C{row_number}"].number_format = '"$"#,##0.00'
        worksheet[f"D{row_number}"].number_format = "0.0%"
        worksheet[f"E{row_number}"].number_format = "@"
    worksheet.add_table(_table("MixedTypesTable", "A1:G7"))
    workbook.save(path)
    workbook.close()
    repack_deterministic(path)
    return path


def _generate_f14(output_dir: Path) -> Path:
    path = output_dir / "sparse.xlsx"
    workbook, _ = _new_workbook("EmptyBefore")
    lone_cells = workbook.create_sheet("LoneCells")
    workbook.create_sheet("EmptyAfter")
    lone_cells["B2"] = 1
    lone_cells["X100"] = 2
    workbook.save(path)
    workbook.close()
    repack_deterministic(path)
    return path


def _generate_f15(output_dir: Path) -> Path:
    path = output_dir / "threeD_ref.xlsx"
    workbook, january = _new_workbook("Jan")
    february = workbook.create_sheet("Feb")
    march = workbook.create_sheet("Mar")
    summary = workbook.create_sheet("Summary")

    for worksheet, value in (
        (january, 10),
        (february, 20),
        (march, 30),
    ):
        worksheet["A1"] = "Metric"
        worksheet["B1"] = "Value"
        worksheet["A2"] = "Revenue"
        worksheet["B2"] = value

    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A2"] = "Quarter revenue"
    summary["B2"] = "=SUM(Jan:Mar!B2)"

    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "Summary", {"B2": CachedValue(60)})
    repack_deterministic(path)
    return path


def _generate_f20(output_dir: Path) -> Path:
    path = output_dir / "stress_map.xlsx"
    workbook, stress = _new_workbook("Stress01")
    for sheet_number in range(2, 41):
        workbook.create_sheet(f"Stress{sheet_number:02d}")

    for block_number in range(12):
        first_column = 1 + block_number * 4
        second_column = first_column + 1
        max_row = 14 - block_number
        stress.cell(row=1, column=first_column, value=f"Block{block_number + 1:02d}")
        stress.cell(row=1, column=second_column, value=f"Value{block_number + 1:02d}")
        for row_number in range(2, max_row + 1):
            stress.cell(
                row=row_number,
                column=first_column,
                value=f"K{block_number + 1:02d}-{row_number - 1:02d}",
            )
            stress.cell(
                row=row_number,
                column=second_column,
                value=(block_number + 1) * 100 + row_number - 1,
            )
        _style_header_cells(
            stress,
            (
                f"{get_column_letter(first_column)}1",
                f"{get_column_letter(second_column)}1",
            ),
        )

    workbook["Stress39"].sheet_state = "hidden"
    workbook["Stress40"].sheet_state = "veryHidden"
    for number in range(1, 61):
        suffixes_and_refs = (
            ("R", "'Stress01'!$A$2:$B$3"),
            ("M", "'Stress01'!$A$2:$A$3,'Stress01'!$E$2:$E$3"),
            ("C", f"={number}"),
            ("F", f"=SUM('Stress01'!$B$2:$B$3)+{number}"),
            ("L", f"=_xlfn.LAMBDA(_xlpm.x,_xlpm.x+{number})"),
        )
        for suffix, refers_to in suffixes_and_refs:
            workbook.defined_names.add(DefinedName(f"N{number:03d}{suffix}", attr_text=refers_to))

    workbook.save(path)
    workbook.close()
    repack_deterministic(path)
    return path


def _generate_f19(output_dir: Path) -> Path:
    path = output_dir / "modern_functions.xlsx"
    workbook, worksheet = _new_workbook("Modern")

    lookup_rows = (
        ("Key", "Value"),
        ("alpha", 10),
        ("beta", 20),
        ("gamma", 30),
    )
    for row_number, values in enumerate(lookup_rows, start=1):
        for column_number, value in enumerate(values, start=8):
            worksheet.cell(row=row_number, column=column_number, value=value)

    # A1 is the dynamic-array anchor and A2 is the saved spill follower value.
    # The formulas deliberately retain Excel's stored modern-function prefixes.
    worksheet["A1"] = "=_xlfn._xlws.FILTER(I2:I4,I2:I4>=20)"
    worksheet["A2"] = 30
    worksheet["B1"] = "=SUM(A1#)"
    worksheet["C1"] = "=SUM(FilteredValues#)"
    worksheet["D1"] = "=_xlfn.LET(_xlpm.rate,I2,_xlpm.bonus,1,_xlpm.rate*3+_xlpm.bonus)"
    worksheet["E1"] = "=DoubleIt(I3)"
    worksheet["F1"] = '=_xlfn.XLOOKUP("beta",H2:H4,I2:I4,"missing")'
    worksheet["G2"] = "=@I2:I4"

    workbook.defined_names.add(
        DefinedName("DoubleIt", attr_text="=_xlfn.LAMBDA(_xlpm.x,_xlpm.x*2)")
    )
    workbook.defined_names.add(DefinedName("FilteredValues", attr_text="'Modern'!$A$1"))

    workbook.save(path)
    workbook.close()
    inject_cached_values(
        path,
        "Modern",
        {
            "A1": CachedValue(20),
            "B1": CachedValue(50),
            "C1": CachedValue(50),
            "D1": CachedValue(31),
            "E1": CachedValue(40),
            "F1": CachedValue(20),
            "G2": CachedValue(10),
        },
    )
    repack_deterministic(path)
    return path


def _generate_f16(output_dir: Path) -> Path:
    path = output_dir / "macro_book.xlsm"
    workbook, worksheet = _new_workbook("MacroModel")
    worksheet.append(("Input", "Doubled"))
    worksheet.append((21, "=A2*2"))
    worksheet["Y1"] = "Stamp target"
    workbook.save(path)
    workbook.close()
    inject_cached_values(path, "MacroModel", {"B2": CachedValue(42)})
    inject_vba_project(path)
    repack_deterministic(path)
    return path


def _generate_f21(output_dir: Path) -> Path:
    path = output_dir / "chart_image.xlsx"
    image_path = output_dir / ".f21-source.png"
    image_path.write_bytes(_PNG_16X16)
    try:
        workbook, worksheet = _new_workbook("Dashboard")
        worksheet.append(("Category", "Value"))
        for category, value in (("North", 10), ("South", 20), ("West", 15), ("East", 25)):
            worksheet.append((category, value))

        chart = BarChart()
        chart.title = "Regional values"
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Category"
        chart.add_data(Reference(worksheet, min_col=2, min_row=1, max_row=5), titles_from_data=True)
        chart.set_categories(Reference(worksheet, min_col=1, min_row=2, max_row=5))
        chart.height = 7.5
        chart.width = 12
        worksheet.add_chart(chart, "D2")

        image = Image(image_path)
        image.width = 32
        image.height = 32
        worksheet.add_image(image, "D17")
        workbook.save(path)
        workbook.close()
    finally:
        image_path.unlink(missing_ok=True)
    repack_deterministic(path)
    return path


def generate_all(output_dir: Path = GENERATED_DIR) -> dict[str, Path]:
    """Generate the implemented deterministic fixtures keyed by fixture ID."""
    output_dir.mkdir(parents=True, exist_ok=True)
    # Keep the original P1 generation order intact so adding fixtures cannot
    # perturb their byte-for-byte golden archives through library global state.
    f01 = _generate_f01(output_dir)
    f07 = _generate_f07(output_dir)
    existing = {
        "F01": f01,
        "F02": _generate_f02(output_dir),
        "F03": _generate_f03(output_dir),
        "F04": _generate_f04(output_dir),
        "F05": _generate_f05(output_dir),
        "F07": f07,
        "F09a": _generate_f09a(output_dir),
        "F09b": _generate_f09b(output_dir),
        "F12": _generate_f12(output_dir),
        "F13": _generate_f13(output_dir),
        "F14": _generate_f14(output_dir),
        "F15": _generate_f15(output_dir),
        "F20": _generate_f20(output_dir),
        "F19": _generate_f19(output_dir),
    }
    # New phase fixtures are authored only after the historical corpus so their
    # library activity cannot perturb the frozen hashes above.
    existing.update(
        {
            "F08": _generate_f08(output_dir),
            "F10": _generate_f10(output_dir),
            "F11": _generate_f11(output_dir),
            "F18": _generate_f18(output_dir),
        }
    )
    existing.update(
        {
            "F16": _generate_f16(output_dir),
            "F21": _generate_f21(output_dir),
        }
    )
    return existing


def main() -> None:
    """Generate fixtures into ``tests/fixtures/generated``."""
    for fixture_id, path in generate_all().items():
        print(f"{fixture_id}: {path}")


if __name__ == "__main__":
    main()
