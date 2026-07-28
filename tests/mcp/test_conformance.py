"""Subprocess stdio conformance for the frozen 14-tool MCP surface."""

from __future__ import annotations

import asyncio
import json
import os
import shutil
import sys
from pathlib import Path
from typing import Any, cast

from mcp import ClientSession, StdioServerParameters
from mcp.client.stdio import stdio_client
from openpyxl import load_workbook

from excel_lsp.core.errors import ErrorCode
from excel_lsp.server.service import RESPONSE_CHARACTER_CAP

FIXTURES = Path(__file__).parents[1] / "fixtures" / "generated"
SCHEMA_SNAPSHOT = Path(__file__).parents[2] / "docs" / "evidence" / "p7-tool-schemas.json"
EXPECTED_TOOLS = {
    "open_workbook",
    "refresh",
    "list_symbols",
    "get_region_schema",
    "read_range",
    "find",
    "trace_precedents",
    "trace_dependents",
    "trace_path",
    "explain_formula",
    "get_diagnostics",
    "profile_column",
    "write_cells",
    "set_column_formula",
}
WRITE_TOOLS = {"write_cells", "set_column_formula"}


def _payload(result: Any) -> dict[str, Any]:
    if result.structuredContent is not None:
        return cast(dict[str, Any], result.structuredContent)
    assert result.content and result.content[0].type == "text"
    return cast(dict[str, Any], json.loads(result.content[0].text))


def _size(payload: dict[str, Any]) -> int:
    return len(json.dumps(payload, ensure_ascii=False, indent=2, default=str))


async def _exercise_server(root: Path, workbook: Path) -> None:
    environment = os.environ.copy()
    environment["EXCEL_LSP_ROOT"] = str(root)
    parameters = StdioServerParameters(
        command=sys.executable,
        args=["-m", "excel_lsp.server"],
        cwd=Path.cwd(),
        env=environment,
    )
    progress_events: list[tuple[float, float | None, str | None]] = []

    async def on_progress(progress: float, total: float | None, message: str | None) -> None:
        progress_events.append((progress, total, message))

    async with stdio_client(parameters) as (read_stream, write_stream):  # noqa: SIM117
        async with ClientSession(read_stream, write_stream) as session:
            initialized = await session.initialize()
            assert initialized.instructions
            assert "open_workbook" in initialized.instructions
            assert "read_range" in initialized.instructions

            listed = await session.list_tools()
            assert {tool.name for tool in listed.tools} == EXPECTED_TOOLS
            expected_schemas = json.loads(SCHEMA_SNAPSHOT.read_text(encoding="utf-8"))
            actual_schemas = {
                tool.name: {
                    "inputSchema": tool.inputSchema,
                    "annotations": tool.annotations.model_dump(by_alias=True, exclude_none=True)
                    if tool.annotations is not None
                    else None,
                }
                for tool in listed.tools
            }
            assert actual_schemas == expected_schemas
            for tool in listed.tools:
                assert tool.description
                assert tool.inputSchema.get("type") == "object"
                assert tool.annotations is not None
                if tool.name in WRITE_TOOLS:
                    assert tool.annotations.readOnlyHint is False
                    assert tool.annotations.destructiveHint is True
                else:
                    assert tool.annotations.readOnlyHint is True
                    assert tool.annotations.openWorldHint is False

            async def call(name: str, arguments: dict[str, Any]) -> dict[str, Any]:
                result = await session.call_tool(name, arguments)
                if result.content and result.content[0].type == "text":
                    assert len(result.content[0].text) <= RESPONSE_CHARACTER_CAP
                payload = _payload(result)
                assert _size(payload) <= RESPONSE_CHARACTER_CAP
                return payload

            opened = await session.call_tool(
                "open_workbook",
                {"path": str(workbook)},
                progress_callback=on_progress,
            )
            opened_payload = _payload(opened)
            assert opened.content[0].type == "text"
            assert len(opened.content[0].text) <= RESPONSE_CHARACTER_CAP
            assert opened_payload["workbook"] == workbook.name
            assert progress_events and progress_events[-1][0] == 3

            happy_calls = (
                ("refresh", {"path": str(workbook), "recalculated": False}),
                ("list_symbols", {"path": str(workbook), "kinds": ["regions"]}),
                (
                    "get_region_schema",
                    {"path": str(workbook), "region_id": "region:Inputs:0"},
                ),
                (
                    "find",
                    {"path": str(workbook), "pattern": "Revenue", "in": ["headers"]},
                ),
                (
                    "trace_precedents",
                    {"path": str(workbook), "ref_or_symbol": "Calc!C2"},
                ),
                (
                    "trace_dependents",
                    {"path": str(workbook), "ref_or_symbol": "Inputs!B2"},
                ),
                (
                    "trace_path",
                    {
                        "path": str(workbook),
                        "from_ref_or_symbol": "Inputs!B2",
                        "to_ref_or_symbol": "Summary!C2",
                    },
                ),
                ("explain_formula", {"path": str(workbook), "ref": "Calc!C2"}),
                ("get_diagnostics", {"path": str(workbook)}),
                ("profile_column", {"path": str(workbook), "col_symbol_or_ref": "Inputs!B:B"}),
            )
            for name, arguments in happy_calls:
                assert "error" not in await call(name, arguments)

            oversized_error = await call(
                "get_region_schema",
                {"path": str(workbook), "region_id": "region:" + "x" * 12_000},
            )
            assert oversized_error["error"]["code"] == ErrorCode.UNKNOWN_SYMBOL.value
            assert oversized_error["truncated"] is True

            first = await call(
                "read_range",
                {"path": str(workbook), "ref": "Inputs!A1:C5", "max_cells": 2},
            )
            assert first["cursor"] and first["truncated"] is True
            second = await call(
                "read_range",
                {
                    "path": str(workbook),
                    "ref": "Inputs!A1:C5",
                    "max_cells": 2,
                    "cursor": first["cursor"],
                },
            )
            assert second["offset"] == 2

            written = await call(
                "write_cells",
                {"path": str(workbook), "cells": [{"ref": "Inputs!B2", "value": 7}]},
            )
            assert written["results"] == [{"ref": "Inputs!B2", "ok": True}]
            assert written["resultsTotal"] == 1
            stale_value_match = await call(
                "find",
                {
                    "path": str(workbook),
                    "pattern": "^600$",
                    "in": ["values"],
                    "sheet": "Calc",
                },
            )
            assert stale_value_match["matches"][0]["ref"] == "cell:Calc!C2"
            assert stale_value_match["stale"] is True
            stale = await call(
                "read_range",
                {
                    "path": str(workbook),
                    "ref": "Inputs!A1:C5",
                    "max_cells": 2,
                    "cursor": first["cursor"],
                },
            )
            assert stale["error"]["code"] == ErrorCode.STALE_CURSOR.value

            column_write = await call(
                "set_column_formula",
                {
                    "path": str(workbook),
                    "col_symbol": "col:Calc:0:cost",
                    "formula": "=RC[-1]*0.5",
                    "overwrite": True,
                },
            )
            assert column_write["cellsWritten"] == 5

            externally_edited = load_workbook(workbook)
            inputs = externally_edited["Inputs"]
            inputs["B2"] = 0.42
            late = externally_edited.create_sheet("Late")
            for row in range(1, 1_101):
                late.cell(row, 1, row)
            sparse = externally_edited.create_sheet("Sparse")
            sparse["A1"] = "Value"
            sparse["A2"] = 2
            sparse["A4"] = 4
            escaped = externally_edited.create_sheet("Escaped")
            escaped["A1"] = "\n" * 4_000
            externally_edited.save(workbook)
            externally_edited.close()
            refreshed_by_read = await call(
                "list_symbols", {"path": str(workbook), "kinds": ["regions"]}
            )
            assert refreshed_by_read["reindexed"] is True
            late_symbol = await call(
                "list_symbols",
                {"path": str(workbook), "query": "A1100", "kinds": ["cells"]},
            )
            assert [item["id"] for item in late_symbol["symbols"]] == ["cell:Late!A1100"]
            assert late_symbol["total"] == 1
            sparse_profile = await call(
                "profile_column",
                {"path": str(workbook), "col_symbol_or_ref": "col:Sparse:0:value"},
            )
            assert sparse_profile["range"] == "A2:A4"
            assert sparse_profile["count"] == 3
            assert sparse_profile["nonnull"] == 2
            escaped_value = await call("read_range", {"path": str(workbook), "ref": "Escaped!A1"})
            assert escaped_value["valueTruncated"] is True
            assert escaped_value["truncated"] is False
            assert escaped_value["cursor"] is None
            returned = escaped_value["values"][0][0]
            assert isinstance(returned, str)
            assert returned.endswith("…")
            assert len(returned) < 4_000
            assert set(returned[:-1]) == {"\n"}

            missing = root / "missing.xlsx"
            outside = root.parent / "outside.xlsx"
            shutil.copyfile(FIXTURES / "basic_single_table.xlsx", outside)
            error_calls = (
                ("open_workbook", {"path": str(missing)}, ErrorCode.NOT_FOUND.value),
                ("refresh", {"path": str(missing)}, ErrorCode.NOT_FOUND.value),
                (
                    "list_symbols",
                    {"path": str(workbook), "kinds": ["bogus"]},
                    ErrorCode.INVALID_VALUE.value,
                ),
                (
                    "get_region_schema",
                    {"path": str(workbook), "region_id": "region:Missing:0"},
                    ErrorCode.UNKNOWN_SYMBOL.value,
                ),
                (
                    "read_range",
                    {"path": str(workbook), "ref": "Inputs!not-a-ref"},
                    ErrorCode.INVALID_REF.value,
                ),
                (
                    "find",
                    {"path": str(workbook), "pattern": "["},
                    ErrorCode.INVALID_VALUE.value,
                ),
                (
                    "trace_precedents",
                    {"path": str(workbook), "ref_or_symbol": "missing"},
                    ErrorCode.INVALID_REF.value,
                ),
                (
                    "trace_dependents",
                    {"path": str(workbook), "ref_or_symbol": "missing"},
                    ErrorCode.INVALID_REF.value,
                ),
                (
                    "trace_path",
                    {
                        "path": str(workbook),
                        "from_ref_or_symbol": "missing",
                        "to_ref_or_symbol": "Inputs!B2",
                    },
                    ErrorCode.INVALID_REF.value,
                ),
                (
                    "explain_formula",
                    {"path": str(workbook), "ref": "Inputs!A1"},
                    ErrorCode.INVALID_REF.value,
                ),
                (
                    "get_diagnostics",
                    {"path": str(workbook), "severity": "critical"},
                    ErrorCode.INVALID_VALUE.value,
                ),
                (
                    "profile_column",
                    {"path": str(workbook), "col_symbol_or_ref": "Inputs!A1:B2"},
                    ErrorCode.INVALID_REF.value,
                ),
                (
                    "set_column_formula",
                    {
                        "path": str(workbook),
                        "col_symbol": "col:Missing:0:x",
                        "formula": "=1",
                    },
                    ErrorCode.UNKNOWN_SYMBOL.value,
                ),
                (
                    "open_workbook",
                    {"path": str(outside)},
                    ErrorCode.PATH_DENIED.value,
                ),
            )
            for name, arguments, expected_code in error_calls:
                payload = await call(name, arguments)
                assert payload["error"]["code"] == expected_code, (name, payload)

            invalid_write = await call(
                "write_cells",
                {"path": str(workbook), "cells": [{"ref": "Inputs!bad", "value": 1}]},
            )
            assert invalid_write["results"][0]["ok"] is False
            assert invalid_write["results"][0]["error"]["code"] == ErrorCode.INVALID_REF.value


def test_stdio_server_conformance(tmp_path: Path) -> None:
    root = tmp_path / "allowed"
    root.mkdir()
    workbook = root / "cross_sheet_model.xlsx"
    shutil.copyfile(FIXTURES / workbook.name, workbook)

    asyncio.run(_exercise_server(root, workbook))
