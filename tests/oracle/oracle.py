"""Canonical openpyxl and production-parser cell streams for T-oracle."""

from __future__ import annotations

from pathlib import Path
from typing import TypeAlias

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from excel_lsp.core.values import JsonScalar, normalize_value

CanonicalCell: TypeAlias = tuple[str, str, JsonScalar, str | None]


def openpyxl_canonical_cells(path: Path) -> tuple[CanonicalCell, ...]:
    """Dual-load a workbook and emit canonical non-empty/formula cell tuples."""
    formula_workbook = load_workbook(path, read_only=True, data_only=False)
    value_workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if formula_workbook.sheetnames != value_workbook.sheetnames:
            raise AssertionError("dual-load sheet order differs")

        canonical: list[CanonicalCell] = []
        for sheet_name in formula_workbook.sheetnames:
            formula_sheet = formula_workbook[sheet_name]
            value_sheet = value_workbook[sheet_name]
            max_row = max(formula_sheet.max_row, value_sheet.max_row)
            max_column = max(formula_sheet.max_column, value_sheet.max_column)
            formula_rows = formula_sheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
            )
            value_rows = value_sheet.iter_rows(
                min_row=1,
                max_row=max_row,
                min_col=1,
                max_col=max_column,
            )
            for row_number, (formula_row, value_row) in enumerate(
                zip(formula_rows, value_rows, strict=True),
                start=1,
            ):
                for column_number, (formula_cell, value_cell) in enumerate(
                    zip(formula_row, value_row, strict=True),
                    start=1,
                ):
                    formula: str | None = None
                    if formula_cell.data_type == "f":
                        if not isinstance(formula_cell.value, str):
                            raise AssertionError(
                                "openpyxl returned a non-string formula for "
                                f"{sheet_name}!{get_column_letter(column_number)}{row_number}"
                            )
                        formula = formula_cell.value
                        if not formula.startswith("="):
                            formula = f"={formula}"

                    raw_value = value_cell.value
                    if formula is None and raw_value is None and formula_cell.value is None:
                        continue
                    ref = f"{get_column_letter(column_number)}{row_number}"
                    canonical.append(
                        (sheet_name, ref, normalize_value(raw_value), formula),
                    )
        return tuple(canonical)
    finally:
        formula_workbook.close()
        value_workbook.close()


def ooxml_canonical_cells(path: Path) -> tuple[CanonicalCell, ...]:
    """Adapt the production OOXMLParser stream to the oracle tuple shape."""
    from excel_lsp.core.parse import OOXMLParser

    canonical: list[CanonicalCell] = []
    with OOXMLParser(path) as parser:
        for descriptor in parser.metadata.sheets:
            for cell in parser.collect_cells(descriptor):
                canonical.append(
                    (
                        descriptor.name,
                        cell.ref,
                        normalize_value(cell.value),
                        cell.formula,
                    )
                )
    return tuple(canonical)
