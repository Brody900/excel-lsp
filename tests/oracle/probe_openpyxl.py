"""Executable probe for pinned openpyxl read-only worksheet behavior."""

from __future__ import annotations

import argparse
import json
import runpy
from collections.abc import Callable
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, cast

import openpyxl
from openpyxl import load_workbook

_FOLLOWER_REFS = ("C3", "C11", "C14", "C21")
_CACHE_REFS = ("C2", "C3", "C11", "C12", "C13", "C14", "C21")
GenerateAll = Callable[[Path], dict[str, Path]]


def _read_only_tables(worksheet: Any) -> dict[str, object]:
    try:
        tables = worksheet.tables
    except AttributeError as error:
        return {
            "available": False,
            "error_type": type(error).__name__,
        }
    return {
        "available": True,
        "names": sorted(tables.keys()),
    }


def _read_only_merges(worksheet: Any) -> dict[str, object]:
    try:
        merged_cells = worksheet.merged_cells
    except AttributeError as error:
        return {
            "available": False,
            "error_type": type(error).__name__,
        }
    return {
        "available": True,
        "ranges": sorted(str(merged_range) for merged_range in merged_cells.ranges),
    }


def probe_read_only(path: Path) -> dict[str, object]:
    """Return JSON-serializable observations for F07 under openpyxl 3.1.5."""
    formula_workbook = load_workbook(path, read_only=True, data_only=False)
    value_workbook = load_workbook(path, read_only=True, data_only=True)
    normal_workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        formula_sheet = formula_workbook["FormulaBlocks"]
        value_sheet = value_workbook["FormulaBlocks"]
        normal_sheet = normal_workbook["FormulaBlocks"]
        return {
            "openpyxl_version": openpyxl.__version__,
            "worksheet_type": type(formula_sheet).__name__,
            "shared_formula_followers": {ref: formula_sheet[ref].value for ref in _FOLLOWER_REFS},
            "formula_caches": {ref: value_sheet[ref].value for ref in _CACHE_REFS},
            "read_only_tables": _read_only_tables(formula_sheet),
            "read_only_merged_cells": _read_only_merges(formula_sheet),
            "normal_mode_control": {
                "table_names": sorted(normal_sheet.tables.keys()),
                "merged_ranges": sorted(
                    str(merged_range) for merged_range in normal_sheet.merged_cells.ranges
                ),
            },
        }
    finally:
        formula_workbook.close()
        value_workbook.close()
        normal_workbook.close()


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        help="existing F07 workbook; omitted to generate one in a temporary directory",
    )
    return parser.parse_args()


def main() -> None:
    """Run the probe and print stable JSON."""
    args = _parse_args()
    if args.workbook is not None:
        result = probe_read_only(args.workbook)
    else:
        generate_all = cast(
            GenerateAll,
            runpy.run_path(str(Path(__file__).parents[1] / "fixtures" / "generate.py"))[
                "generate_all"
            ],
        )

        with TemporaryDirectory(prefix="excel-lsp-openpyxl-probe-") as directory:
            result = probe_read_only(generate_all(Path(directory))["F07"])
    print(json.dumps(result, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
