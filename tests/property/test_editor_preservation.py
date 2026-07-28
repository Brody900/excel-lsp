from __future__ import annotations

from collections.abc import Sequence
from pathlib import Path
from zipfile import ZipFile

from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st
from openpyxl import Workbook

from excel_lsp.core.edit import CellEdit, CellEditKind, WriteScalar, patch_workbook
from excel_lsp.core.models import CellRecord
from excel_lsp.core.parse import OOXMLParser


@st.composite
def _edit_scripts(draw: st.DrawFn) -> tuple[CellEdit, ...]:
    coordinates = draw(
        st.lists(
            st.tuples(st.integers(1, 8), st.integers(1, 5)),
            min_size=1,
            max_size=10,
            unique=True,
        )
    )
    edits: list[CellEdit] = []
    value_strategy = st.one_of(
        st.none(),
        st.booleans(),
        st.integers(-10_000, 10_000),
        st.floats(-10_000, 10_000, allow_nan=False, allow_infinity=False),
        st.text(
            alphabet=st.characters(
                blacklist_categories=("Cc", "Cs"),
                blacklist_characters=("\ufffe", "\uffff"),
            ),
            max_size=30,
        ),
    )
    formula_strategy = st.sampled_from(("=1+2", "=A1+1", "=SUM(A1:B2)"))
    for row, col in coordinates:
        ref = f"{'ABCDE'[col - 1]}{row}"
        if draw(st.booleans()):
            edits.append(CellEdit.formula("Property", ref, draw(formula_strategy)))
        else:
            edits.append(CellEdit.value("Property", ref, draw(value_strategy)))
    return tuple(edits)


def _author_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Property"
    for row in range(1, 6):
        worksheet.cell(row=row, column=1, value=row)
        worksheet.cell(row=row, column=2, value=row * 2)
        worksheet.cell(row=row, column=3, value=f"=A{row}+B{row}")
    workbook.save(path)
    workbook.close()


def _parts(path: Path) -> dict[str, bytes]:
    with ZipFile(path) as archive:
        return {
            info.filename: archive.read(info) for info in archive.infolist() if not info.is_dir()
        }


def _parsed_cells(path: Path) -> dict[str, CellRecord]:
    with OOXMLParser(path) as parser:
        descriptor = parser.metadata.sheets[0]
        cells: dict[str, CellRecord] = {}
        parser.parse_sheet(descriptor, lambda cell: cells.__setitem__(cell.ref, cell))
        return cells


def _expected_value(value: WriteScalar) -> WriteScalar:
    if type(value) is float and value.is_integer():
        return int(value)
    return value


@given(_edit_scripts())
@settings(
    max_examples=50,
    deadline=None,
    suppress_health_check=(HealthCheck.function_scoped_fixture,),
)
def test_i18_random_edit_scripts_preserve_parts_and_reparse_exactly(
    tmp_path: Path,
    edits: Sequence[CellEdit],
) -> None:
    workbook = tmp_path / "property.xlsx"
    _author_workbook(workbook)
    before = _parts(workbook)
    with OOXMLParser(workbook) as parser:
        expected_hash = parser.hashes.whole_file

    result = patch_workbook(workbook, edits, expected_workbook_hash=expected_hash)

    after = _parts(workbook)
    assert set(after) == set(before) - set(result.deleted_parts)
    actual_modified = {part for part in before.keys() & after.keys() if before[part] != after[part]}
    assert actual_modified == set(result.modified_parts)
    for part, payload in before.items():
        if part not in result.modified_parts and part not in result.deleted_parts:
            assert after[part] == payload
    parsed = _parsed_cells(workbook)
    for edit in edits:
        if edit.kind is CellEditKind.FORMULA:
            assert parsed[edit.ref].formula == edit.payload
            assert parsed[edit.ref].value is None
        elif edit.payload is None:
            assert edit.ref not in parsed
        else:
            assert parsed[edit.ref].formula is None
            assert parsed[edit.ref].value == _expected_value(edit.payload)
