"""P8 exact-answer and deterministic scripted benchmark regressions."""

from __future__ import annotations

import asyncio
import csv
import json
from pathlib import Path
from zipfile import ZipFile

import lxml.etree as etree
import pytest
from openpyxl import load_workbook

from benchmarks.baseline_server import server as baseline_server
from benchmarks.check import AnswerContractError, check_transcript, parse_final_answer
from benchmarks.model import TASKS
from benchmarks.run_scripted import collect_scripted, write_scripted
from benchmarks.workloads import ARCHIVE_ROWS, build_workloads

_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _existing_cell_semantics(payload: bytes) -> dict[str, object]:
    root = etree.fromstring(payload)
    result: dict[str, object] = {}
    for cell in root.findall(f".//{{{_MAIN_NS}}}c"):
        ref = cell.get("r")
        row = int("".join(character for character in ref if character.isdigit())) if ref else 0
        if ref is None or row >= 1_000:
            continue
        result[ref] = (
            tuple(sorted(cell.attrib.items())),
            tuple(
                (
                    etree.QName(child).localname,
                    tuple(sorted(child.attrib.items())),
                    tuple(child.itertext()),
                )
                for child in cell
            ),
        )
    return result


def test_every_task_has_frozen_prompt_answer_shape_and_exact_answer() -> None:
    root = Path(__file__).parents[2]

    assert [task.task_id for task in TASKS] == [f"B{number}" for number in range(1, 7)]
    for task in TASKS:
        prompt_path = root / "benchmarks" / "tasks" / f"{task.task_id}.md"
        prompt = prompt_path.read_text(encoding="utf-8")
        assert prompt == task.markdown()
        generated = task.prompt(
            root / "tests" / "fixtures" / "generated" / "benchmarks" / task.fixture
        )
        assert generated.endswith("The last line of your reply must be exactly: `ANSWER: <json>`")
        transcript = "Reasoning may appear here.\nANSWER: " + json.dumps(
            task.expected, ensure_ascii=False, separators=(",", ":")
        )
        passed, reason = check_transcript(task.task_id, transcript)
        assert passed is True
        assert reason == ("exact set" if task.unordered_array_key else "exact")


@pytest.mark.parametrize(
    "transcript",
    (
        "",
        'ANSWER:{"value":1464.1}',
        'ANSWER: {"value":1464.1}\ntrailing prose',
        'ANSWER: {"value":1464.1} ',
        "ANSWER: not-json",
    ),
)
def test_answer_parser_rejects_non_exact_final_line(transcript: str) -> None:
    with pytest.raises(AnswerContractError):
        parse_final_answer(transcript)


def test_checker_accepts_set_order_but_rejects_duplicates_shape_and_unknown_task() -> None:
    reordered = 'ANSWER: {"input_cells":["Inputs!B3","Inputs!B2","Inputs!B4","Inputs!B5"]}'
    duplicate = 'ANSWER: {"input_cells":["Inputs!B2","Inputs!B2","Inputs!B4","Inputs!B5"]}'
    assert check_transcript("B1", reordered) == (True, "exact set")
    assert check_transcript("B1", duplicate)[0] is False
    assert check_transcript("B1", 'ANSWER: {"cells":[]}')[0] is False
    assert check_transcript("B99", "ANSWER: {}") == (False, "unknown task: B99")


def test_naive_baseline_tools_are_explicitly_read_only() -> None:
    tools = asyncio.run(baseline_server.list_tools())

    assert {tool.name for tool in tools} == {"read_workbook_full", "read_sheet"}
    assert all(tool.annotations is not None for tool in tools)
    assert all(tool.annotations.readOnlyHint is True for tool in tools if tool.annotations)
    assert all(tool.annotations.openWorldHint is False for tool in tools if tool.annotations)


def test_scripted_replays_cover_both_arms_and_write_recheckable_rows(tmp_path: Path) -> None:
    root = Path(__file__).parents[2]
    results = collect_scripted(root)

    assert {(row.task, row.arm) for row in results} == {
        (f"B{number}", arm) for number in range(1, 7) for arm in ("excel-lsp", "naive-dump")
    }
    assert all(row.status == "ok" and row.correct for row in results)
    assert all(
        row.payload_tokens > 0 and row.tool_calls > 0 and row.wall_ms >= 0 for row in results
    )
    for task in TASKS:
        excel = next(row for row in results if row.task == task.task_id and row.arm == "excel-lsp")
        baseline = next(
            row for row in results if row.task == task.task_id and row.arm == "naive-dump"
        )
        assert check_transcript(task.task_id, excel.transcript)[0] is True
        assert check_transcript(task.task_id, baseline.transcript)[0] is True
        assert baseline.payload_tokens >= excel.payload_tokens * 10

    output = tmp_path / "scripted.csv"
    write_scripted(results, output)
    with output.open(encoding="utf-8", newline="") as stream:
        rows = list(csv.DictReader(stream))
    assert len(rows) == 12
    assert {row["correct"] for row in rows} == {"True"}


def test_benchmark_workloads_are_deterministic_disclosed_archive_envelopes(
    tmp_path: Path,
) -> None:
    root = Path(__file__).parents[2]
    first = build_workloads(root, output_dir=tmp_path / "first", force=True)
    second = build_workloads(root, output_dir=tmp_path / "second", force=True)

    assert first.keys() == second.keys()
    for fixture in first:
        workbook = load_workbook(first[fixture], read_only=True, data_only=False)
        repeated = load_workbook(second[fixture], read_only=True, data_only=False)
        try:
            assert workbook["BenchmarkArchive"].max_row == ARCHIVE_ROWS + 1
            assert tuple(workbook["BenchmarkArchive"].values) == tuple(
                repeated["BenchmarkArchive"].values
            )
            if fixture == "cross_sheet_model.xlsx":
                assert workbook["Summary"].max_row == 2_000
        finally:
            workbook.close()
            repeated.close()

        source = root / "tests" / "fixtures" / "generated" / fixture
        with ZipFile(source) as source_archive, ZipFile(first[fixture]) as workload_archive:
            source_members = {name: source_archive.read(name) for name in source_archive.namelist()}
            workload_members = {
                name: workload_archive.read(name) for name in workload_archive.namelist()
            }
        deliberately_modified = {
            "[Content_Types].xml",
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
        }
        if fixture == "cross_sheet_model.xlsx":
            deliberately_modified.add("xl/worksheets/sheet3.xml")
        assert {
            name: payload
            for name, payload in workload_members.items()
            if name in source_members and name not in deliberately_modified
        } == {
            name: payload
            for name, payload in source_members.items()
            if name not in deliberately_modified
        }

        if fixture == "cross_sheet_model.xlsx":
            assert _existing_cell_semantics(
                source_members["xl/worksheets/sheet3.xml"]
            ) == _existing_cell_semantics(workload_members["xl/worksheets/sheet3.xml"])
