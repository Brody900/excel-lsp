"""Real-parser integration checks for the Phase 1 lifecycle."""

from __future__ import annotations

import hashlib
import os
import runpy
from collections.abc import Callable
from pathlib import Path
from typing import cast
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from excel_lsp.core.index import IndexStore, index_workbook
from excel_lsp.core.parse import OOXMLParser

GenerateAll = Callable[[Path], dict[str, Path]]

generate_all = cast(
    GenerateAll,
    runpy.run_path(str(Path(__file__).parents[1] / "fixtures" / "generate.py"))["generate_all"],
)


def _rewrite_zip_member(path: Path, member: str, payload: bytes) -> None:
    temporary = path.with_name(f"{path.name}.tmp")
    with ZipFile(path) as source, ZipFile(temporary, mode="w") as destination:
        found = False
        for info in source.infolist():
            data = payload if info.filename == member else source.read(info.filename)
            found = found or info.filename == member
            destination.writestr(info, data)
    assert found
    temporary.replace(path)


def test_real_parser_indexes_f01_and_then_takes_the_noop_path(tmp_path: Path) -> None:
    workbook = generate_all(tmp_path / "fixtures")["F01"]
    source_hash = hashlib.sha256(workbook.read_bytes()).hexdigest()

    first = index_workbook(workbook, index_dir=tmp_path / "indexes")

    assert first.changed is True
    assert first.reindexed_sheets == ("Sales",)
    with IndexStore(first.index_path) as store:
        assert store.connection.execute("SELECT COUNT(*) FROM cells").fetchone()[0] == 24
        assert tuple(
            store.connection.execute(
                """
                SELECT value, value_type, formula
                FROM cells
                WHERE ref = 'D3'
                """
            ).fetchone()
        ) == (6.25, "number", "=B3*C3")
        assert store.get_meta("has_vba") == "0"
        assert store.get_meta("external_links") == "{}"

    second = index_workbook(workbook, index_dir=tmp_path / "indexes")

    assert second.changed is False
    assert second.generation == first.generation
    assert second.reindexed_sheets == ()
    assert hashlib.sha256(workbook.read_bytes()).hexdigest() == source_hash


def test_mtime_only_change_updates_source_stat_without_bumping_generation(tmp_path: Path) -> None:
    workbook = generate_all(tmp_path / "fixtures")["F01"]
    source_bytes = workbook.read_bytes()
    first = index_workbook(workbook, index_dir=tmp_path / "indexes")
    original_stat = workbook.stat()
    touched_mtime_ns = original_stat.st_mtime_ns + 2_000_000_000

    os.utime(workbook, ns=(original_stat.st_atime_ns, touched_mtime_ns))
    touched_stat = workbook.stat()
    assert touched_stat.st_mtime_ns != original_stat.st_mtime_ns
    assert workbook.read_bytes() == source_bytes

    refreshed = index_workbook(workbook, index_dir=tmp_path / "indexes")

    assert refreshed.changed is False
    assert refreshed.generation == first.generation
    assert refreshed.reindexed_sheets == ()
    with IndexStore(refreshed.index_path) as store:
        assert store.generation == first.generation
        assert store.get_meta("mtime_ns") == str(touched_stat.st_mtime_ns)
        assert store.get_meta("size") == str(touched_stat.st_size)

    fast_noop = index_workbook(workbook, index_dir=tmp_path / "indexes")
    assert fast_noop.changed is False
    assert fast_noop.generation == first.generation
    assert fast_noop.reindexed_sheets == ()
    assert workbook.read_bytes() == source_bytes


@pytest.mark.parametrize("mutation", ["name", "ref"])
def test_table_part_only_change_reindexes_the_owning_sheet(
    tmp_path: Path,
    mutation: str,
) -> None:
    workbook_path = tmp_path / f"table-{mutation}.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Data"
    worksheet.append(("Value",))
    worksheet.append((1,))
    worksheet.append((2,))
    worksheet.add_table(Table(displayName="Table1", ref="A1:A2"))
    workbook.save(workbook_path)
    workbook.close()

    first = index_workbook(workbook_path, index_dir=tmp_path / f"index-{mutation}")
    with IndexStore(first.index_path) as store:
        old_table_hash = store.get_part_hash("xl/tables/table1.xml")
        assert old_table_hash is not None
        assert store.get_part_hash("xl/worksheets/_rels/sheet1.xml.rels") is not None
    with ZipFile(workbook_path) as archive:
        old_sheet_xml = archive.read("xl/worksheets/sheet1.xml")
        old_table_xml = archive.read("xl/tables/table1.xml")
    old_mtime = workbook_path.stat().st_mtime_ns

    workbook = load_workbook(workbook_path)
    worksheet = workbook["Data"]
    table = next(iter(worksheet.tables.values()))
    if mutation == "name":
        table.name = "RenamedTable"
        table.displayName = "RenamedTable"
        expected_name = "RenamedTable"
        expected_ref = "A1:A2"
    else:
        table.ref = "A1:A3"
        expected_name = "Table1"
        expected_ref = "A1:A3"
    workbook.save(workbook_path)
    workbook.close()
    if workbook_path.stat().st_mtime_ns <= old_mtime:
        next_mtime = old_mtime + 1_000_000
        os.utime(workbook_path, ns=(next_mtime, next_mtime))

    with ZipFile(workbook_path) as archive:
        assert archive.read("xl/worksheets/sheet1.xml") == old_sheet_xml
        assert archive.read("xl/tables/table1.xml") != old_table_xml

    update = index_workbook(workbook_path, index_dir=tmp_path / f"index-{mutation}")

    assert update.reindexed_sheets == ("Data",)
    with IndexStore(update.index_path) as store:
        assert store.get_part_hash("xl/tables/table1.xml") != old_table_hash
    with OOXMLParser(workbook_path) as parser:
        summary = parser.parse_sheet(parser.metadata.sheets[0], lambda _cell: None)
    assert [(table.name, table.ref) for table in summary.tables] == [(expected_name, expected_ref)]


def test_content_type_only_sheet_reclassification_rebuilds_the_catalog(tmp_path: Path) -> None:
    workbook_path = generate_all(tmp_path / "fixtures")["F01"]
    first = index_workbook(workbook_path, index_dir=tmp_path / "index")
    with IndexStore(first.index_path) as store:
        old_content_type_hash = store.get_part_hash("[Content_Types].xml")
        assert old_content_type_hash is not None
    old_mtime = workbook_path.stat().st_mtime_ns
    with ZipFile(workbook_path) as archive:
        content_types = archive.read("[Content_Types].xml")
    worksheet_type = b"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
    chartsheet_type = b"application/vnd.openxmlformats-officedocument.spreadsheetml.chartsheet+xml"
    assert worksheet_type in content_types
    _rewrite_zip_member(
        workbook_path,
        "[Content_Types].xml",
        content_types.replace(worksheet_type, chartsheet_type, 1),
    )
    if workbook_path.stat().st_mtime_ns <= old_mtime:
        next_mtime = old_mtime + 1_000_000
        os.utime(workbook_path, ns=(next_mtime, next_mtime))

    update = index_workbook(workbook_path, index_dir=tmp_path / "index")

    assert update.reindexed_sheets == ("Sales",)
    with IndexStore(update.index_path) as store:
        assert tuple(store.connection.execute("SELECT kind FROM sheets").fetchone()) == (
            "chartsheet",
        )
        assert store.connection.execute("SELECT COUNT(*) FROM cells").fetchone()[0] == 0
        assert store.get_part_hash("[Content_Types].xml") != old_content_type_hash
