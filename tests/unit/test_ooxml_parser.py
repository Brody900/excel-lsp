"""High-signal contract tests for the streaming OOXML package parser."""

from __future__ import annotations

import hashlib
from collections.abc import Mapping
from datetime import datetime
from pathlib import Path
from textwrap import dedent
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.formula import DataTableFormula

import excel_lsp.core.parse.parser as parser_module
from excel_lsp.core.errors import ErrorCode, ExcelLSPError
from excel_lsp.core.index import IndexStore
from excel_lsp.core.models import (
    CalculationProperties,
    CellRecord,
    DataTableFormulaInfo,
    DataValidationInfo,
    NameArea,
    Rect,
    TableInfo,
)
from excel_lsp.core.parse import OOXMLParser

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
REL_TYPE_BASE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _xml(source: str) -> bytes:
    return dedent(source).strip().encode()


def _write_package(path: Path, parts: Mapping[str, bytes]) -> Path:
    with ZipFile(path, mode="w", compression=ZIP_DEFLATED) as archive:
        for name, data in parts.items():
            archive.writestr(name, data)
    return path


def _contract_parts() -> dict[str, bytes]:
    return {
        "[Content_Types].xml": _xml(
            f"""
            <Types xmlns="{CONTENT_TYPES_NS}">
              <Override PartName="/xl/workbook.xml"
                ContentType="application/vnd.ms-excel.sheet.macroEnabled.main+xml"/>
              <Override PartName="/xl/worksheets/sheet1.xml"
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
              <Override PartName="/xl/worksheets/sheet2.xml"
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
              <Override PartName="/xl/chartsheets/sheet1.xml"
                ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.chartsheet+xml"/>
            </Types>
            """
        ),
        "xl/workbook.xml": _xml(
            f"""
            <workbook xmlns="{MAIN_NS}" xmlns:r="{DOCUMENT_REL_NS}">
              <workbookPr date1904="true"/>
              <sheets>
                <sheet name="Data" sheetId="1" r:id="rIdData"/>
                <sheet name="NoDimension" sheetId="2" state="hidden" r:id="rIdNoDim"/>
                <sheet name="Dashboard" sheetId="3" state="veryHidden" r:id="rIdChart"/>
              </sheets>
              <definedNames>
                <definedName name="GlobalRange">'Data'!$A$1:$B$3</definedName>
                <definedName name="LocalRange" localSheetId="1">$B$7</definedName>
                <definedName name="MultiArea">'Data'!$A$1:$A$2,'NoDimension'!$B$7:$B$8</definedName>
                <definedName name="Constant">=42</definedName>
                <definedName name="Formula">=SUM('Data'!$C$1:$C$3)</definedName>
                <definedName name="Increment">=_xlfn.LAMBDA(x,x+1)</definedName>
                <definedName name="RelativeLocal" localSheetId="1">B7</definedName>
                <definedName name="MixedGlobal">'Data'!$A1</definedName>
                <definedName name="AbsoluteColumns">'Data'!$A:$B</definedName>
                <definedName name="MixedColumns">'Data'!$A:B</definedName>
                <definedName name="AbsoluteRows">'Data'!$1:$2</definedName>
                <definedName name="MixedRows">'Data'!$1:2</definedName>
                <definedName name="_xlnm.Print_Area">'Data'!$A$1:$D$3</definedName>
              </definedNames>
              <externalReferences>
                <externalReference r:id="rIdExternal"/>
              </externalReferences>
              <calcPr calcId="191029" calcMode="manual" fullCalcOnLoad="1"
                refMode="R1C1" iterate="true" iterateCount="42" iterateDelta="0.0001"
                fullPrecision="0" calcCompleted="1" calcOnSave="false"
                concurrentCalc="1" concurrentManualCount="3" forceFullCalc="true"/>
            </workbook>
            """
        ),
        "xl/_rels/workbook.xml.rels": _xml(
            f"""
            <Relationships xmlns="{PACKAGE_REL_NS}">
              <Relationship Id="rIdData" Type="{REL_TYPE_BASE}/worksheet"
                Target="worksheets/sheet1.xml"/>
              <Relationship Id="rIdNoDim" Type="{REL_TYPE_BASE}/worksheet"
                Target="worksheets/sheet2.xml"/>
              <Relationship Id="rIdChart" Type="{REL_TYPE_BASE}/chartsheet"
                Target="chartsheets/sheet1.xml"/>
              <Relationship Id="rIdExternal" Type="{REL_TYPE_BASE}/externalLink"
                Target="externalLinks/externalLink1.xml"/>
            </Relationships>
            """
        ),
        "xl/sharedStrings.xml": _xml(
            f"""
            <sst xmlns="{MAIN_NS}" count="2" uniqueCount="2">
              <si>
                <r><t xml:space="preserve"> rich </t></r>
                <r><t>text</t></r>
                <rPh><t>ignored phonetic text</t></rPh>
              </si>
              <si><t>plain</t></si>
            </sst>
            """
        ),
        "xl/styles.xml": _xml(
            f"""
            <styleSheet xmlns="{MAIN_NS}">
              <numFmts count="1">
                <numFmt numFmtId="164" formatCode="yyyy-mm-dd"/>
              </numFmts>
              <fonts count="1"><font><b/></font></fonts>
              <fills count="1">
                <fill><patternFill patternType="solid"><fgColor rgb="FFFF00"/></patternFill></fill>
              </fills>
              <cellXfs count="3">
                <xf numFmtId="0" fontId="0" fillId="0"/>
                <xf numFmtId="14" fontId="0" fillId="0"/>
                <xf numFmtId="164" fontId="0" fillId="0"/>
              </cellXfs>
            </styleSheet>
            """
        ),
        "xl/worksheets/sheet1.xml": _xml(
            f"""
            <worksheet xmlns="{MAIN_NS}" xmlns:r="{DOCUMENT_REL_NS}">
              <dimension ref="A1:A1"/>
              <sheetData>
                <row r="1">
                  <c r="A1" t="s"><v>0</v></c>
                  <c r="B1" t="inlineStr">
                    <is><r><t xml:space="preserve">inline </t></r><r><t>value</t></r></is>
                  </c>
                  <c r="C1"><v>42</v></c>
                  <c r="D1" t="n"><v>3.5</v></c>
                  <c r="E1" t="b"><v>1</v></c>
                  <c r="F1" t="b"><v>false</v></c>
                  <c r="G1" t="e"><v>#DIV/0!</v></c>
                  <c r="H1" t="str"><f>TEXT(C1,"0")</f><v>42 items</v></c>
                  <c r="I1" s="1"><v>1.5</v></c>
                  <c r="J1" s="2"><v>2</v></c>
                </row>
                <row r="2">
                  <c r="A2"><v>2</v></c>
                  <c r="B2"><v>10</v></c>
                  <c r="C2"><f t="shared" si="0" ref="C2:C3">A2*B2</f><v>20</v></c>
                  <c r="D2"><f t="array" ref="D2:D3">ROW(A1:A2)</f><v>1</v></c>
                </row>
                <row r="3">
                  <c r="A3"><v>3</v></c>
                  <c r="B3"><v>10</v></c>
                  <c r="C3"><f t="shared" si="0"/><v>30</v></c>
                  <c r="D3"><v>2</v></c>
                </row>
                <row r="42"><c r="Z42"><v>9</v></c></row>
              </sheetData>
              <mergeCells count="1"><mergeCell ref="E5:F5"/></mergeCells>
              <dataValidations count="1">
                <dataValidation type="whole" operator="between" allowBlank="1"
                  sqref="A2:A3 C2:D3">
                  <formula1>1</formula1><formula2>10</formula2>
                </dataValidation>
              </dataValidations>
              <tableParts count="1"><tablePart r:id="rIdTable"/></tableParts>
            </worksheet>
            """
        ),
        # A deliberately non-standard namespace verifies local-name matching.
        "xl/worksheets/sheet2.xml": _xml(
            """
            <producer:worksheet xmlns:producer="urn:producer-specific-spreadsheetml">
              <producer:sheetData>
                <producer:row r="7">
                  <producer:c r="B7"><producer:v>77</producer:v></producer:c>
                </producer:row>
              </producer:sheetData>
            </producer:worksheet>
            """
        ),
        # Cell-looking content must not be parsed when the workbook classifies this as a chartsheet.
        "xl/chartsheets/sheet1.xml": _xml(
            f"""
            <chartsheet xmlns="{MAIN_NS}">
              <sheetData><row r="1"><c r="XFD1048576"><v>999</v></c></row></sheetData>
            </chartsheet>
            """
        ),
        "xl/worksheets/_rels/sheet1.xml.rels": _xml(
            f"""
            <Relationships xmlns="{PACKAGE_REL_NS}">
              <Relationship Id="rIdTable" Type="{REL_TYPE_BASE}/table"
                Target="../tables/table1.xml"/>
            </Relationships>
            """
        ),
        "xl/tables/table1.xml": _xml(
            f"""
            <table xmlns="{MAIN_NS}" name="Orders" displayName="Orders" ref="A1:D3"
              headerRowCount="1" totalsRowCount="1">
              <tableColumns count="4">
                <tableColumn id="1" name="Item"/>
                <tableColumn id="2" name="Quantity"/>
                <tableColumn id="3" name="Price"/>
                <tableColumn id="4" name="Total"/>
              </tableColumns>
            </table>
            """
        ),
        "xl/externalLinks/externalLink1.xml": _xml(
            f"""
            <externalLink xmlns="{MAIN_NS}" xmlns:r="{DOCUMENT_REL_NS}">
              <externalBook r:id="rIdBook"/>
            </externalLink>
            """
        ),
        "xl/externalLinks/_rels/externalLink1.xml.rels": _xml(
            f"""
            <Relationships xmlns="{PACKAGE_REL_NS}">
              <Relationship Id="rIdBook" Type="{REL_TYPE_BASE}/externalLinkPath"
                Target="https://example.test/source.xlsx" TargetMode="External"/>
            </Relationships>
            """
        ),
        "xl/vbaProject.bin": b"test-only-vba-marker",
    }


def _contract_package(tmp_path: Path) -> tuple[Path, dict[str, bytes]]:
    parts = _contract_parts()
    return _write_package(tmp_path / "contract.xlsm", parts), parts


def _windows_date_parts() -> dict[str, bytes]:
    return {
        "xl/workbook.xml": _xml(
            f"""
            <workbook xmlns="{MAIN_NS}" xmlns:r="{DOCUMENT_REL_NS}">
              <workbookPr date1904="0"/>
              <sheets><sheet name="Dates" sheetId="1" r:id="rId1"/></sheets>
            </workbook>
            """
        ),
        "xl/_rels/workbook.xml.rels": _xml(
            f"""
            <Relationships xmlns="{PACKAGE_REL_NS}">
              <Relationship Id="rId1" Type="{REL_TYPE_BASE}/worksheet"
                Target="worksheets/sheet1.xml"/>
            </Relationships>
            """
        ),
        "xl/styles.xml": _xml(
            f"""
            <styleSheet xmlns="{MAIN_NS}">
              <cellXfs count="2"><xf numFmtId="0"/><xf numFmtId="14"/></cellXfs>
            </styleSheet>
            """
        ),
        "xl/worksheets/sheet1.xml": _xml(
            f"""
            <worksheet xmlns="{MAIN_NS}"><sheetData><row r="1">
              <c r="A1" s="1"><v>59</v></c>
              <c r="B1" s="1"><v>60</v></c>
              <c r="C1" s="1"><v>61</v></c>
            </row></sheetData></worksheet>
            """
        ),
    }


def test_streams_all_cell_types_formulas_and_actual_bounds(tmp_path: Path) -> None:
    path, _ = _contract_package(tmp_path)
    streamed: list[CellRecord] = []

    with OOXMLParser(path) as parser:
        data_sheet, no_dimension_sheet, _ = parser.metadata.sheets
        summary = parser.parse_sheet(data_sheet, streamed.append)
        no_dimension_cells: list[CellRecord] = []
        no_dimension_summary = parser.parse_sheet(no_dimension_sheet, no_dimension_cells.append)

        assert parser.metadata.date1904 is True
        assert parser.styles.custom_num_formats == {164: "yyyy-mm-dd"}
        assert parser.styles.fonts[0].bold is True
        assert parser.styles.fills[0].pattern_type == "solid"

    by_ref = {cell.ref: cell for cell in streamed}
    assert [(by_ref[ref].value, by_ref[ref].value_type) for ref in ("A1", "B1")] == [
        (" rich text", "string"),
        ("inline value", "string"),
    ]
    assert [(by_ref[ref].value, by_ref[ref].value_type) for ref in ("C1", "D1")] == [
        (42, "number"),
        (3.5, "number"),
    ]
    assert [(by_ref[ref].value, by_ref[ref].value_type) for ref in ("E1", "F1")] == [
        (True, "bool"),
        (False, "bool"),
    ]
    assert (by_ref["G1"].value, by_ref["G1"].value_type) == ("#DIV/0!", "error")
    assert (by_ref["H1"].value, by_ref["H1"].value_type, by_ref["H1"].formula) == (
        "42 items",
        "string",
        '=TEXT(C1,"0")',
    )
    assert (by_ref["I1"].value, by_ref["I1"].value_type) == (
        datetime(1904, 1, 2, 12),
        "date",
    )
    assert (by_ref["J1"].value, by_ref["J1"].value_type) == (
        datetime(1904, 1, 3),
        "date",
    )

    assert (by_ref["C2"].formula, by_ref["C2"].formula_kind, by_ref["C2"].shared_index) == (
        "=A2*B2",
        "shared",
        0,
    )
    assert (by_ref["C3"].formula, by_ref["C3"].formula_kind, by_ref["C3"].shared_index) == (
        "=A3*B3",
        "shared",
        0,
    )
    assert (by_ref["D2"].formula, by_ref["D2"].formula_kind, by_ref["D2"].array_ref) == (
        "=ROW(A1:A2)",
        "array",
        "D2:D3",
    )
    assert (by_ref["D3"].formula, by_ref["D3"].formula_kind, by_ref["D3"].array_ref) == (
        None,
        "array",
        "D2:D3",
    )

    assert [cell.ref for cell in streamed][-1] == "Z42"
    assert (summary.dimension_ref, summary.max_row, summary.max_col, summary.cell_count) == (
        "A1:A1",
        42,
        26,
        len(streamed),
    )
    assert no_dimension_cells == [CellRecord(ref="B7", row=7, col=2, value=77, value_type="number")]
    assert (
        no_dimension_summary.dimension_ref,
        no_dimension_summary.max_row,
        no_dimension_summary.max_col,
    ) == (None, 7, 2)


@pytest.mark.parametrize(
    "value",
    ("3.5", "1.0", "1e2", "9007199254740993.0", "1e309", "1e-9999"),
)
def test_plain_numeric_fast_parser_preserves_canonical_number_semantics(value: str) -> None:
    fast = parser_module._parse_plain_number(value)
    canonical = parser_module._parse_number(value)

    assert fast == canonical
    assert type(fast) is type(canonical)


def test_collects_array_merges_multirange_validation_and_listobject(tmp_path: Path) -> None:
    path, _ = _contract_package(tmp_path)

    with OOXMLParser(path) as parser:
        summary = parser.parse_sheet(parser.metadata.sheets[0], lambda _cell: None)

    assert summary.array_formulas == (Rect(2, 3, 4, 4),)
    assert summary.merges == (Rect(5, 5, 5, 6),)
    assert summary.validations == (
        DataValidationInfo(Rect(2, 3, 1, 1), "whole", "between", "1", "10", True),
        DataValidationInfo(Rect(2, 3, 3, 4), "whole", "between", "1", "10", True),
    )
    assert summary.tables == (
        TableInfo(
            name="Orders",
            display_name="Orders",
            ref="A1:D3",
            header_rows=1,
            totals_rows=1,
            columns=("Item", "Quantity", "Price", "Total"),
        ),
    )


def test_parses_names_external_links_vba_and_tolerates_chartsheet(tmp_path: Path) -> None:
    path, _ = _contract_package(tmp_path)

    with OOXMLParser(path) as parser:
        metadata = parser.metadata
        chart_summary = parser.parse_sheet(metadata.sheets[2], lambda _cell: pytest.fail())

    assert [(sheet.name, sheet.kind, sheet.visibility) for sheet in metadata.sheets] == [
        ("Data", "worksheet", "visible"),
        ("NoDimension", "worksheet", "hidden"),
        ("Dashboard", "chartsheet", "veryHidden"),
    ]
    assert (chart_summary.cell_count, chart_summary.max_row, chart_summary.max_col) == (0, 0, 0)
    assert metadata.external_links == {1: "https://example.test/source.xlsx"}
    assert metadata.has_vba is True
    assert metadata.calculation == CalculationProperties(
        calc_id=191029,
        calc_mode="manual",
        full_calc_on_load=True,
        ref_mode="R1C1",
        iterate=True,
        iterate_count=42,
        iterate_delta=0.0001,
        full_precision=False,
        calc_completed=True,
        calc_on_save=False,
        concurrent_calc=True,
        concurrent_manual_count=3,
        force_full_calc=True,
    )

    names = {name.name: name for name in metadata.defined_names}
    assert (names["GlobalRange"].scope_sheet_order, names["GlobalRange"].kind) == (
        None,
        "range",
    )
    assert names["GlobalRange"].areas == (NameArea("Data", Rect(1, 3, 1, 2)),)
    assert (names["LocalRange"].scope_sheet_order, names["LocalRange"].kind) == (1, "range")
    assert names["LocalRange"].areas == (NameArea("NoDimension", Rect(7, 7, 2, 2)),)
    assert names["MultiArea"].kind == "multi_range"
    assert names["MultiArea"].areas == (
        NameArea("Data", Rect(1, 2, 1, 1)),
        NameArea("NoDimension", Rect(7, 8, 2, 2)),
    )
    assert names["Constant"].kind == "constant"
    assert names["Formula"].kind == "formula"
    assert names["Increment"].kind == "lambda"
    assert (names["RelativeLocal"].kind, names["RelativeLocal"].areas) == ("formula", ())
    assert (names["MixedGlobal"].kind, names["MixedGlobal"].areas) == ("formula", ())
    assert names["AbsoluteColumns"].areas == (NameArea("Data", Rect(1, 1_048_576, 1, 2)),)
    assert (names["MixedColumns"].kind, names["MixedColumns"].areas) == ("formula", ())
    assert names["AbsoluteRows"].areas == (NameArea("Data", Rect(1, 2, 1, 16_384)),)
    assert (names["MixedRows"].kind, names["MixedRows"].areas) == ("formula", ())
    assert names["_xlnm.Print_Area"].is_builtin is True


def test_hashes_selected_parts_and_never_mutates_source(tmp_path: Path) -> None:
    path, parts = _contract_package(tmp_path)
    before = path.read_bytes()

    with OOXMLParser(path) as parser:
        hashes = parser.hashes
        for sheet in parser.metadata.sheets:
            parser.parse_sheet(sheet, lambda _cell: None)

    selected_parts = {
        "[Content_Types].xml",
        "xl/workbook.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/sharedStrings.xml",
        "xl/styles.xml",
        "xl/worksheets/sheet1.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/tables/table1.xml",
        "xl/worksheets/sheet2.xml",
        "xl/chartsheets/sheet1.xml",
        "xl/externalLinks/externalLink1.xml",
        "xl/externalLinks/_rels/externalLink1.xml.rels",
    }
    assert hashes.whole_file == hashlib.sha256(before).hexdigest()
    assert hashes.parts.keys() == selected_parts
    assert hashes.parts == {
        name: hashlib.sha256(parts[name]).hexdigest() for name in selected_parts
    }
    assert path.read_bytes() == before


def test_duplicate_sheet_names_are_reported_as_corrupt_ooxml(tmp_path: Path) -> None:
    parts = _contract_parts()
    parts["xl/workbook.xml"] = _xml(
        f"""
        <workbook xmlns="{MAIN_NS}" xmlns:r="{DOCUMENT_REL_NS}">
          <sheets>
            <sheet name="Data" sheetId="1" r:id="rIdData"/>
            <sheet name="data" sheetId="2" r:id="rIdNoDim"/>
          </sheets>
        </workbook>
        """
    )
    path = _write_package(tmp_path / "duplicate-sheets.xlsx", parts)

    with pytest.raises(ExcelLSPError) as caught, OOXMLParser(path):
        pass

    assert caught.value.code is ErrorCode.CORRUPT


def test_duplicate_case_insensitive_defined_names_are_reported_as_corrupt_ooxml(
    tmp_path: Path,
) -> None:
    parts = _contract_parts()
    parts["xl/workbook.xml"] = parts["xl/workbook.xml"].replace(
        b"</definedNames>",
        b"<definedName name=\"GLOBALRANGE\">'Data'!$A$1</definedName></definedNames>",
    )
    path = _write_package(tmp_path / "duplicate-defined-names.xlsx", parts)

    with pytest.raises(ExcelLSPError) as caught, OOXMLParser(path):
        pass

    assert caught.value.code is ErrorCode.CORRUPT


def test_duplicate_cell_coordinates_are_reported_as_corrupt_ooxml(tmp_path: Path) -> None:
    parts = _contract_parts()
    parts["xl/worksheets/sheet1.xml"] = _xml(
        f"""
        <worksheet xmlns="{MAIN_NS}"><sheetData><row r="1">
          <c r="A1"><v>1</v></c><c r="A1"><v>2</v></c>
        </row></sheetData></worksheet>
        """
    )
    path = _write_package(tmp_path / "duplicate-cells.xlsx", parts)

    with OOXMLParser(path) as parser, pytest.raises(ExcelLSPError) as caught:
        parser.collect_cells(parser.metadata.sheets[0])

    assert caught.value.code is ErrorCode.CORRUPT


@pytest.mark.parametrize(
    "table_xml",
    [
        f"""
        <table xmlns="{MAIN_NS}" name="Orders" displayName="Orders" ref="A1:D3">
          <tableColumns count="3">
            <tableColumn id="1" name="A"/>
            <tableColumn id="2" name="B"/>
            <tableColumn id="3" name="C"/>
          </tableColumns>
        </table>
        """,
        f"""
        <table xmlns="{MAIN_NS}" name="Orders" displayName="Orders" ref="A1:A1"
          headerRowCount="1" totalsRowCount="1">
          <tableColumns count="1"><tableColumn id="1" name="A"/></tableColumns>
        </table>
        """,
    ],
)
def test_invalid_table_geometry_is_reported_as_corrupt_ooxml(
    tmp_path: Path,
    table_xml: str,
) -> None:
    parts = _contract_parts()
    parts["xl/tables/table1.xml"] = _xml(table_xml)
    path = _write_package(tmp_path / "invalid-table.xlsx", parts)

    with OOXMLParser(path) as parser, pytest.raises(ExcelLSPError) as caught:
        parser.collect_cells(parser.metadata.sheets[0])

    assert caught.value.code is ErrorCode.CORRUPT


def test_windows_date_serial_60_uses_openpyxl_reference_conversion(tmp_path: Path) -> None:
    path = _write_package(tmp_path / "windows-dates.xlsx", _windows_date_parts())

    with OOXMLParser(path) as parser:
        assert parser.metadata.date1904 is False
        assert parser.metadata.calculation == CalculationProperties()
        cells = parser.collect_cells(parser.metadata.sheets[0])

    assert [(cell.ref, cell.value, cell.value_type) for cell in cells] == [
        ("A1", datetime(1900, 2, 28), "date"),
        ("B1", datetime(1900, 2, 28), "date"),
        ("C1", datetime(1900, 3, 1), "date"),
    ]


def test_large_integral_excel_number_indexes_as_sqlite_real(tmp_path: Path) -> None:
    parts = _contract_parts()
    parts["xl/worksheets/sheet1.xml"] = _xml(
        f"""
        <worksheet xmlns="{MAIN_NS}">
          <sheetData><row r="1"><c r="A1"><v>1E+20</v></c></row></sheetData>
        </worksheet>
        """
    )
    path = _write_package(tmp_path / "large-integral.xlsx", parts)

    with OOXMLParser(path) as parser, IndexStore(tmp_path / "large-integral.xlsp.db") as store:
        descriptor = parser.metadata.sheets[0]
        store.replace_sheet_catalog(parser.metadata.sheets)
        store.replace_sheet(
            descriptor,
            lambda on_cell: parser.parse_sheet(descriptor, on_cell),
        )
        row = store.connection.execute(
            "SELECT value, typeof(value) FROM cells WHERE ref = 'A1'"
        ).fetchone()

    assert tuple(row) == (1e20, "real")


def test_pinned_openpyxl_data_table_formula_is_parsed_as_typed_opaque_span(
    tmp_path: Path,
) -> None:
    path = tmp_path / "data-table.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["A1"] = 1
    worksheet["B1"] = DataTableFormula(ref="B1:B3", r1="A1", ca=True)
    worksheet["B2"] = 2
    worksheet["B3"] = 3
    workbook.save(path)
    workbook.close()

    with OOXMLParser(path) as parser:
        summary = parser.parse_sheet(parser.metadata.sheets[0], lambda _cell: None)
        cells = parser.collect_cells(parser.metadata.sheets[0])

    assert summary.data_tables == (
        DataTableFormulaInfo(
            ref="B1:B3",
            rect=Rect(1, 3, 2, 2),
            input_cell_1="A1",
            calculate_always=True,
        ),
    )
    by_ref = {cell.ref: cell for cell in cells}
    assert by_ref["B1"].formula is None
    for ref in ("B1", "B2", "B3"):
        assert by_ref[ref].formula_kind == "dataTable"
        assert by_ref[ref].array_ref == "B1:B3"
        assert by_ref[ref].data_table == summary.data_tables[0]


@pytest.mark.parametrize(
    "formula_xml",
    [
        '<c r="C2"><f t="shared" si="0">A2*B2</f><v>2</v></c>',
        (
            '<c r="C2"><f t="shared" si="0" ref="C2:C3">A2*B2</f><v>2</v></c>'
            '<c r="C4"><f t="shared" si="0"/><v>4</v></c>'
        ),
    ],
    ids=["master-missing-ref", "follower-outside-span"],
)
def test_shared_formula_groups_require_and_enforce_master_span(
    tmp_path: Path,
    formula_xml: str,
) -> None:
    parts = _contract_parts()
    parts["xl/worksheets/sheet1.xml"] = _xml(
        f"""
        <worksheet xmlns="{MAIN_NS}">
          <sheetData><row r="2">{formula_xml}</row></sheetData>
        </worksheet>
        """
    )
    path = _write_package(tmp_path / "invalid-shared-formula.xlsx", parts)

    with OOXMLParser(path) as parser, pytest.raises(ExcelLSPError) as caught:
        parser.collect_cells(parser.metadata.sheets[0])

    assert caught.value.code is ErrorCode.CORRUPT


def test_shared_formula_translation_handles_modern_reference_syntax(
    tmp_path: Path,
) -> None:
    parts = _contract_parts()
    parts["xl/worksheets/sheet1.xml"] = _xml(
        f"""
        <worksheet xmlns="{MAIN_NS}">
          <sheetData>
            <row r="2">
              <c r="B2"><f t="shared" si="1" ref="B2:B3">@A2</f><v>1</v></c>
              <c r="C2"><f t="shared" si="2" ref="C2:C3">A2:INDEX(A:A,2)</f><v>1</v></c>
              <c r="D2"><f t="shared" si="3" ref="D2:D3">A2#</f><v>1</v></c>
              <c r="E2"><f t="shared" si="4" ref="E2:E3">SUM(A2#)</f><v>1</v></c>
              <c r="F2"><f t="shared" si="5" ref="F2:F3">Esc[A'[B]</f><v>1</v></c>
            </row>
            <row r="3">
              <c r="B3"><f t="shared" si="1"/><v>1</v></c>
              <c r="C3"><f t="shared" si="2"/><v>1</v></c>
              <c r="D3"><f t="shared" si="3"/><v>1</v></c>
              <c r="E3"><f t="shared" si="4"/><v>1</v></c>
              <c r="F3"><f t="shared" si="5"/><v>1</v></c>
            </row>
          </sheetData>
        </worksheet>
        """
    )
    path = _write_package(tmp_path / "modern-shared-formulas.xlsx", parts)

    with OOXMLParser(path) as parser:
        cells = parser.collect_cells(parser.metadata.sheets[0])

    formulas = {cell.ref: cell.formula for cell in cells}
    assert formulas == {
        "B2": "=@A2",
        "C2": "=A2:INDEX(A:A,2)",
        "D2": "=A2#",
        "E2": "=SUM(A2#)",
        "F2": "=Esc[A'[B]",
        "B3": "=@A3",
        "C3": "=A3:INDEX(A:A,2)",
        "D3": "=A3#",
        "E3": "=SUM(A3#)",
        "F3": "=Esc[A'[B]",
    }


def test_locked_package_open_retries_once_after_the_frozen_delay(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "locked.xlsx"
    path.write_bytes(b"exists so path validation succeeds")
    attempts = 0
    delays: list[float] = []

    def locked_hash(_path: Path) -> tuple[str, bytes]:
        nonlocal attempts
        attempts += 1
        raise PermissionError("sharing violation")

    monkeypatch.setattr(parser_module, "_hash_file", locked_hash)
    monkeypatch.setattr(parser_module.time_module, "sleep", delays.append)

    with pytest.raises(ExcelLSPError) as caught, OOXMLParser(path):
        pass

    assert caught.value.code is ErrorCode.LOCKED
    assert attempts == 2
    assert delays == [0.5]


def test_open_failures_use_structured_error_codes(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    unsupported = tmp_path / "legacy.xls"
    unsupported.write_bytes(b"not relevant: suffix is rejected first")
    encrypted = tmp_path / "encrypted.xlsx"
    encrypted.write_bytes(bytes.fromhex("D0CF11E0A1B11AE1") + b"test payload")
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not a zip package")
    missing_part = _write_package(
        tmp_path / "missing-workbook-part.xlsx",
        {"[Content_Types].xml": _xml(f'<Types xmlns="{CONTENT_TYPES_NS}"/>')},
    )
    monkeypatch.setattr(OOXMLParser, "retry_delay_seconds", 0.0)

    cases = (
        (tmp_path / "missing.xlsx", ErrorCode.NOT_FOUND),
        (unsupported, ErrorCode.UNSUPPORTED_FORMAT),
        (encrypted, ErrorCode.ENCRYPTED),
        (corrupt, ErrorCode.CORRUPT),
        (missing_part, ErrorCode.CORRUPT),
    )
    for path, expected_code in cases:
        with pytest.raises(ExcelLSPError) as caught, OOXMLParser(path):
            pass
        assert caught.value.code is expected_code
        payload = caught.value.as_dict()
        assert payload["error"]["code"] == expected_code.value
        assert payload["error"]["message"]
