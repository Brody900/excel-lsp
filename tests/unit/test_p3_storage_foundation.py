"""P3 storage and parser dependency regressions."""

from __future__ import annotations

import hashlib
import json
import os
import sqlite3
from collections.abc import Callable
from pathlib import Path
from textwrap import dedent
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from excel_lsp.core.errors import ErrorCode, ExcelLSPError
from excel_lsp.core.index.lifecycle import index_workbook
from excel_lsp.core.index.schema import SCHEMA_VERSION
from excel_lsp.core.index.store import IndexStore
from excel_lsp.core.models import CellRecord, SheetDescriptor, SheetParseSummary, TableInfo
from excel_lsp.core.parse import OOXMLParser
from excel_lsp.core.parse.coordinates import make_cell_ref, parse_rect

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOCUMENT_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
REL_TYPE_BASE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _descriptor(name: str, order: int) -> SheetDescriptor:
    return SheetDescriptor(
        order=order,
        name=name,
        sheet_id=order + 1,
        rel_id=f"rId{order + 1}",
        xml_part=f"xl/worksheets/sheet{order + 1}.xml",
        kind="worksheet",
    )


def _replace_sheet(
    store: IndexStore,
    descriptor: SheetDescriptor,
    *tables: TableInfo,
    part_hash: str,
) -> SheetParseSummary:
    cells: list[CellRecord] = []
    for table in tables:
        rect = parse_rect(table.ref)
        for offset, name in enumerate(table.columns):
            column = rect.col_min + offset
            cells.append(
                CellRecord(
                    ref=make_cell_ref(rect.row_min, column),
                    row=rect.row_min,
                    col=column,
                    value=name,
                    value_type="string",
                )
            )

    def parse(on_cell: Callable[[CellRecord], None]) -> SheetParseSummary:
        for cell in cells:
            on_cell(cell)
        return SheetParseSummary(
            descriptor=descriptor,
            part_hash=part_hash,
            max_row=max((parse_rect(table.ref).row_max for table in tables), default=0),
            max_col=max((parse_rect(table.ref).col_max for table in tables), default=0),
            cell_count=len(cells),
            tables=tuple(tables),
        )

    return store.replace_sheet(descriptor, parse)


def _table(
    name: str,
    *,
    ref: str = "B2:D8",
    display_name: str | None = None,
    columns: tuple[str, ...] = ("Item", "Net Sales", "Straße"),
) -> TableInfo:
    return TableInfo(
        name=name,
        display_name=display_name or name,
        ref=ref,
        header_rows=1,
        totals_rows=1,
        columns=columns,
    )


def test_schema_v5_has_normalized_list_object_tables_and_complete_persistence(
    tmp_path: Path,
) -> None:
    descriptor = _descriptor("Data", 0)
    table = _table("SalesTable", display_name="Sales Display")

    with IndexStore(tmp_path / "catalog.xlsp.db") as store:
        assert SCHEMA_VERSION == "5"
        store.replace_sheet_catalog((descriptor,))
        _replace_sheet(store, descriptor, table, part_hash="sheet-v1")

        tables = {
            str(row[0])
            for row in store.connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
        assert {"list_objects", "list_object_columns"} <= tables

        object_columns = {
            str(row[1])
            for row in store.connection.execute("PRAGMA table_info(list_objects)").fetchall()
        }
        child_columns = {
            str(row[1])
            for row in store.connection.execute("PRAGMA table_info(list_object_columns)").fetchall()
        }
        assert {
            "sheet_id",
            "name",
            "lookup_name",
            "display_name",
            "row_min",
            "row_max",
            "col_min",
            "col_max",
            "header_rows",
            "totals_rows",
        } <= object_columns
        assert {"list_object_id", "idx", "name", "lookup_name"} <= child_columns
        assert "totals_rows" not in {
            str(row[1]) for row in store.connection.execute("PRAGMA table_info(regions)").fetchall()
        }

        catalog = store.connection.execute(
            """
            SELECT s.name, t.name, t.lookup_name, t.display_name,
                   t.row_min, t.row_max, t.col_min, t.col_max,
                   t.header_rows, t.totals_rows
            FROM list_objects AS t
            JOIN sheets AS s ON s.id = t.sheet_id
            """
        ).fetchone()
        assert tuple(catalog) == (
            "Data",
            "SalesTable",
            "salestable",
            "Sales Display",
            2,
            8,
            2,
            4,
            1,
            1,
        )
        columns = store.connection.execute(
            """
            SELECT idx, name, lookup_name
            FROM list_object_columns
            ORDER BY idx
            """
        ).fetchall()
        assert tuple(map(tuple, columns)) == (
            (0, "Item", "item"),
            (1, "Net Sales", "net sales"),
            (2, "Straße", "strasse"),
        )
        assert store.canonical_export()["list_object_columns"] == (
            ("Data", "SalesTable", 0, "Item"),
            ("Data", "SalesTable", 1, "Net Sales"),
            ("Data", "SalesTable", 2, "Straße"),
        )


def test_case_insensitive_workbook_duplicate_rolls_back_sheet_replacement(
    tmp_path: Path,
) -> None:
    first = _descriptor("First", 0)
    second = _descriptor("Second", 1)
    with IndexStore(tmp_path / "duplicate.xlsp.db") as store:
        store.replace_sheet_catalog((first, second))
        _replace_sheet(store, first, _table("SalesTable"), part_hash="first-v1")
        _replace_sheet(
            store,
            second,
            _table("InventoryTable", ref="A1:C5"),
            part_hash="second-v1",
        )
        before = store.canonical_export()
        generation = store.generation

        with pytest.raises(ExcelLSPError) as raised:
            _replace_sheet(
                store,
                second,
                _table("SALESTABLE", ref="E2:G7"),
                part_hash="second-invalid",
            )

        assert raised.value.code is ErrorCode.CORRUPT
        assert raised.value.details == {"table": "SALESTABLE", "sheet": "Second"}
        assert store.generation == generation
        assert store.canonical_export() == before


def test_table_display_alias_collision_rolls_back_sheet_replacement(
    tmp_path: Path,
) -> None:
    first = _descriptor("First", 0)
    second = _descriptor("Second", 1)
    with IndexStore(tmp_path / "display-alias-duplicate.xlsp.db") as store:
        store.replace_sheet_catalog((first, second))
        _replace_sheet(
            store,
            first,
            _table("SalesTable", display_name="PublishedSales"),
            part_hash="first-v1",
        )
        _replace_sheet(
            store,
            second,
            _table("InventoryTable", ref="A1:C5"),
            part_hash="second-v1",
        )
        before = store.canonical_export()
        generation = store.generation

        with pytest.raises(ExcelLSPError) as raised:
            _replace_sheet(
                store,
                second,
                _table(
                    "ReplacementTable",
                    ref="E2:G7",
                    display_name="PUBLISHEDSALES",
                ),
                part_hash="second-invalid",
            )

        assert raised.value.code is ErrorCode.CORRUPT
        assert raised.value.details == {"table": "ReplacementTable", "sheet": "Second"}
        assert store.generation == generation
        assert store.canonical_export() == before


def test_table_display_alias_collision_is_structured_and_lifecycle_atomic(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "display-alias-collision.xlsx"
    workbook = Workbook()
    first_sheet = workbook.active
    assert first_sheet is not None
    first_sheet.title = "First"
    first_sheet.append(("Value",))
    first_sheet.append((1,))
    first_sheet.add_table(Table(displayName="TableA", ref="A1:A2"))
    second_sheet = workbook.create_sheet("Second")
    second_sheet.append(("Value",))
    second_sheet.append((2,))
    second_sheet.add_table(Table(displayName="TableB", ref="A1:A2"))
    workbook.save(workbook_path)
    workbook.close()

    index_dir = tmp_path / "indexes"
    initial = index_workbook(workbook_path, index_dir=index_dir)
    with IndexStore(initial.index_path) as store:
        before = store.canonical_export()
        generation = store.generation

    with ZipFile(workbook_path) as archive:
        table_xml = archive.read("xl/tables/table1.xml")
    assert b'displayName="TableA"' in table_xml
    _rewrite_zip_member(
        workbook_path,
        "xl/tables/table1.xml",
        table_xml.replace(b'displayName="TableA"', b'displayName="tableb"'),
    )
    current_stat = workbook_path.stat()
    os.utime(
        workbook_path,
        ns=(current_stat.st_atime_ns, current_stat.st_mtime_ns + 1_000_000),
    )

    with pytest.raises(ExcelLSPError) as raised:
        index_workbook(workbook_path, index_dir=index_dir)

    assert raised.value.code is ErrorCode.CORRUPT
    with IndexStore(initial.index_path) as store:
        assert store.generation == generation
        assert store.canonical_export() == before

    with pytest.raises(ExcelLSPError) as fresh_raised:
        index_workbook(workbook_path, index_dir=tmp_path / "fresh-indexes")
    assert fresh_raised.value.code is ErrorCode.CORRUPT


def test_per_sheet_replacement_replaces_only_its_list_object_catalog(tmp_path: Path) -> None:
    first = _descriptor("First", 0)
    second = _descriptor("Second", 1)
    with IndexStore(tmp_path / "replace.xlsp.db") as store:
        store.replace_sheet_catalog((first, second))
        _replace_sheet(store, first, _table("OldTable"), part_hash="first-v1")
        _replace_sheet(
            store,
            second,
            _table("KeepTable", ref="F2:H6"),
            part_hash="second-v1",
        )
        _replace_sheet(
            store,
            first,
            _table("NewTable", ref="A3:C9", columns=("One", "Two", "Three")),
            part_hash="first-v2",
        )

        objects = store.connection.execute(
            """
            SELECT s.name, t.name, t.row_min, t.row_max, t.col_min, t.col_max
            FROM list_objects AS t
            JOIN sheets AS s ON s.id = t.sheet_id
            ORDER BY s.id
            """
        ).fetchall()
        assert tuple(map(tuple, objects)) == (
            ("First", "NewTable", 3, 9, 1, 3),
            ("Second", "KeepTable", 2, 6, 6, 8),
        )
        child_tables = store.connection.execute(
            """
            SELECT t.name, COUNT(*)
            FROM list_object_columns AS c
            JOIN list_objects AS t ON t.id = c.list_object_id
            GROUP BY t.name
            ORDER BY t.name
            """
        ).fetchall()
        assert tuple(map(tuple, child_tables)) == (("KeepTable", 3), ("NewTable", 3))


def test_incremental_refresh_moves_table_alias_to_an_earlier_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "move-table.xlsx"
    workbook = Workbook()
    first_sheet = workbook.active
    assert first_sheet is not None
    first_sheet.title = "First"
    first_sheet.append(("Value",))
    first_sheet.append((1,))
    second_sheet = workbook.create_sheet("Second")
    second_sheet.append(("Value",))
    second_sheet.append((2,))
    second_sheet.add_table(Table(displayName="MovedTable", ref="A1:A2"))
    workbook.save(workbook_path)
    workbook.close()

    index_dir = tmp_path / "indexes"
    first = index_workbook(workbook_path, index_dir=index_dir)

    moved = load_workbook(workbook_path)
    del moved["Second"].tables["MovedTable"]
    moved["First"].add_table(Table(displayName="MovedTable", ref="A1:A2"))
    moved.save(workbook_path)
    moved.close()

    incremental = index_workbook(workbook_path, index_dir=index_dir)
    fresh = index_workbook(workbook_path, index_dir=tmp_path / "fresh-indexes")

    with IndexStore(incremental.index_path) as incremental_store:
        incremental_export = incremental_store.canonical_export()
    with IndexStore(fresh.index_path) as fresh_store:
        fresh_export = fresh_store.canonical_export()

    assert incremental.generation == first.generation + 1
    assert incremental.reindexed_sheets == ("First", "Second")
    assert incremental_export["list_objects"] == (
        ("First", "MovedTable", "MovedTable", 1, 2, 1, 1, 1, 0),
    )
    assert incremental_export == fresh_export


def test_schema_v2_sidecar_rebuild_preserves_monotonic_generation(tmp_path: Path) -> None:
    database = tmp_path / "schema-v2.xlsp.db"
    connection = sqlite3.connect(database)
    try:
        connection.executescript(
            """
            CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT NOT NULL);
            INSERT INTO meta(key, value) VALUES ('schema_version', '2');
            INSERT INTO meta(key, value) VALUES ('generation', '41');
            CREATE TABLE legacy_v2_marker (value TEXT);
            INSERT INTO legacy_v2_marker(value) VALUES ('drop-me');
            PRAGMA user_version = 2;
            """
        )
        connection.commit()
    finally:
        connection.close()

    with IndexStore(database) as rebuilt:
        assert rebuilt.schema_rebuilt is True
        assert rebuilt.get_meta("schema_version") == "5"
        assert rebuilt.generation == 42
        assert rebuilt.connection.execute("PRAGMA user_version").fetchone()[0] == 5
        tables = {
            str(row[0])
            for row in rebuilt.connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
        assert {"list_objects", "list_object_columns"} <= tables
        assert "legacy_v2_marker" not in tables


def test_external_link_xml_and_relationship_hashes_detect_target_only_change(
    tmp_path: Path,
) -> None:
    path = tmp_path / "external.xlsx"
    first_bytes = _write_external_package(path, "../first-budget.xlsx")
    with OOXMLParser(path) as parser:
        first_hashes = dict(parser.hashes.parts)
        assert parser.metadata.external_links == {1: "../first-budget.xlsx"}

    second_bytes = _write_external_package(path, "../second-budget.xlsx")
    with OOXMLParser(path) as parser:
        second_hashes = dict(parser.hashes.parts)
        assert parser.metadata.external_links == {1: "../second-budget.xlsx"}

    link_part = "xl/externalLinks/externalLink1.xml"
    rels_part = "xl/externalLinks/_rels/externalLink1.xml.rels"
    assert {link_part, rels_part} <= first_hashes.keys() == second_hashes.keys()
    assert first_hashes[link_part] == second_hashes[link_part]
    assert first_hashes[rels_part] != second_hashes[rels_part]
    assert {name for name in first_hashes if first_hashes[name] != second_hashes[name]} == {
        rels_part
    }
    assert first_hashes[link_part] == hashlib.sha256(first_bytes[link_part]).hexdigest()
    assert first_hashes[rels_part] == hashlib.sha256(first_bytes[rels_part]).hexdigest()
    assert second_hashes[rels_part] == hashlib.sha256(second_bytes[rels_part]).hexdigest()


def test_lifecycle_keeps_raw_external_target_private_from_canonical_export(
    tmp_path: Path,
) -> None:
    secret_target = (
        "https://user:password@example.test/private/budget.xlsx?sig=SECRET_TOKEN#fragment"
    )
    workbook = tmp_path / "external-secret.xlsx"
    _write_external_package(workbook, secret_target)

    update = index_workbook(workbook, index_dir=tmp_path / "indexes")

    with IndexStore(update.index_path) as store:
        assert json.loads(store.get_meta("external_links") or "{}") == {"1": secret_target}
        serialized = json.dumps(store.canonical_export(), ensure_ascii=False)

    for secret in ("user", "password", "private", "SECRET_TOKEN", "fragment"):
        assert secret not in serialized


def _write_external_package(path: Path, target: str) -> dict[str, bytes]:
    parts = _external_parts(target)
    with ZipFile(path, mode="w", compression=ZIP_DEFLATED) as archive:
        for name, data in parts.items():
            archive.writestr(name, data)
    return parts


def _rewrite_zip_member(path: Path, member: str, payload: bytes) -> None:
    temporary = path.with_name(f"{path.name}.tmp")
    with ZipFile(path) as source, ZipFile(temporary, mode="w") as destination:
        found = False
        for info in source.infolist():
            data = payload if info.filename == member else source.read(info.filename)
            destination.writestr(info, data)
            found = found or info.filename == member
    if not found:
        raise AssertionError(f"archive member is missing: {member}")
    temporary.replace(path)


def _external_parts(target: str) -> dict[str, bytes]:
    return {
        "xl/workbook.xml": _xml(
            f"""
            <workbook xmlns="{MAIN_NS}" xmlns:r="{DOCUMENT_REL_NS}">
              <sheets/>
              <externalReferences>
                <externalReference r:id="rIdExternal"/>
              </externalReferences>
            </workbook>
            """
        ),
        "xl/_rels/workbook.xml.rels": _xml(
            f"""
            <Relationships xmlns="{PACKAGE_REL_NS}">
              <Relationship Id="rIdExternal" Type="{REL_TYPE_BASE}/externalLink"
                Target="externalLinks/externalLink1.xml"/>
            </Relationships>
            """
        ),
        "xl/externalLinks/externalLink1.xml": _xml(
            f"""
            <externalLink xmlns="{MAIN_NS}" xmlns:r="{DOCUMENT_REL_NS}">
              <externalBook r:id="rIdBook"/>
            </externalLink>
            """
        ),
        "xl/externalLinks/_rels/externalLink1.xml.rels": _xml(
            f"""
            <Relationships xmlns="{PACKAGE_REL_NS}">
              <Relationship Id="rIdBook" Type="{REL_TYPE_BASE}/externalLinkPath"
                Target="{target}" TargetMode="External"/>
            </Relationships>
            """
        ),
    }


def _xml(source: str) -> bytes:
    return dedent(source).strip().encode("utf-8")
