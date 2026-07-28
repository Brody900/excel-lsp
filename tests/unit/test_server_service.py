"""Focused P7 service, cursor, cap, and confinement regressions."""

from __future__ import annotations

import asyncio
import json
import shutil
import time
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excel_lsp.core.errors import ErrorCode, ExcelLSPError
from excel_lsp.server import app as server_app
from excel_lsp.server import service as service_module
from excel_lsp.server.cursors import decode_cursor, encode_cursor, parameter_hash
from excel_lsp.server.models import WriteCellInput
from excel_lsp.server.service import RESPONSE_CHARACTER_CAP, ToolService

FIXTURES = Path(__file__).parents[1] / "fixtures" / "generated"


def _copy_fixture(tmp_path: Path, name: str) -> Path:
    destination = tmp_path / name
    shutil.copyfile(FIXTURES / name, destination)
    return destination


def _serialized(payload: object) -> str:
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"), default=str)


def test_cursor_is_opaque_and_bound_to_tool_parameters_and_generation() -> None:
    digest = parameter_hash({"path": "book.xlsx", "ref": "Sheet1!A1:B2", "max_cells": 2})
    cursor = encode_cursor(tool="read_range", params_hash=digest, offset=2, generation=7)

    assert decode_cursor(cursor, tool="read_range", params_hash=digest, generation=7) == 2
    for changed in (
        {"tool": "find", "params_hash": digest, "generation": 7},
        {"tool": "read_range", "params_hash": "different", "generation": 7},
        {"tool": "read_range", "params_hash": digest, "generation": 8},
    ):
        with pytest.raises(ExcelLSPError) as caught:
            decode_cursor(cursor, **changed)
        assert caught.value.code is ErrorCode.STALE_CURSOR


def test_read_range_pages_sparse_values_and_write_invalidates_cursor(tmp_path: Path) -> None:
    workbook = _copy_fixture(tmp_path, "basic_single_table.xlsx")
    service = ToolService()

    first = service.read_range(str(workbook), "Sales!A1:D3", max_cells=3)
    assert first["truncated"] is True
    assert sum(len(row) for row in first["values"]) == 3
    assert len(_serialized(first)) <= RESPONSE_CHARACTER_CAP
    cursor = first["cursor"]
    assert isinstance(cursor, str)

    second = service.read_range(str(workbook), "Sales!A1:D3", cursor=cursor, max_cells=3)
    assert second["offset"] == 3
    assert second["range"] == "A1:D3"
    assert second["page"] == {"start": "D1", "end": "B2"}
    assert sum(len(row) for row in second["values"]) == 3

    written = service.write_cells(
        str(workbook),
        [WriteCellInput.model_validate({"ref": "Sales!B2", "value": 99})],
    )
    assert written["results"] == [{"ref": "Sales!B2", "ok": True}]

    with pytest.raises(ExcelLSPError) as caught:
        service.read_range(str(workbook), "Sales!A1:D3", cursor=cursor, max_cells=3)
    assert caught.value.code is ErrorCode.STALE_CURSOR


def test_read_range_truncates_json_expanding_single_value_without_pagination(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Escaped"
    original = "\n" * 4_000
    sheet["A1"] = original
    path = tmp_path / "escaped-value.xlsx"
    workbook.save(path)
    service = ToolService()

    first = service.read_range(str(path), "Escaped!A1")
    second = service.read_range(str(path), "Escaped!A1")

    assert len(json.dumps(first, ensure_ascii=False, indent=2)) <= RESPONSE_CHARACTER_CAP
    assert first["valueTruncated"] is True
    assert first["truncated"] is False
    assert first["cursor"] is None
    assert first["values"] == second["values"]
    returned = first["values"][0][0]
    assert isinstance(returned, str)
    assert returned.endswith("…")
    assert len(returned) < len(original)
    assert original.startswith(returned[:-1])


def test_all_read_service_responses_are_bounded() -> None:
    workbook = FIXTURES / "cross_sheet_model.xlsx"
    service = ToolService()
    responses = (
        service.open_workbook(str(workbook)),
        service.refresh(str(workbook)),
        service.list_symbols(str(workbook)),
        service.get_region_schema(str(workbook), "region:Inputs:0"),
        service.read_range(str(workbook), "Inputs!A1:C5"),
        service.find(str(workbook), "Revenue"),
        service.trace_precedents(str(workbook), "Calc!C2"),
        service.trace_dependents(str(workbook), "Inputs!B2"),
        service.trace_path(str(workbook), "Inputs!B2", "Summary!C2"),
        service.explain_formula(str(workbook), "Calc!C2"),
        service.get_diagnostics(str(workbook)),
        service.profile_column(str(workbook), "Inputs!B:B"),
        service.trace_dependents(str(workbook), "sheet:Inputs", depth=1),
        service.trace_precedents(str(workbook), "fblock:Calc:2", depth=0),
    )

    assert all(len(_serialized(response)) <= RESPONSE_CHARACTER_CAP for response in responses)
    profile = responses[11]
    assert profile["range"] == "B2:B5"
    assert profile["sum"] == pytest.approx(1000.95)
    assert responses[-1]["tree"]["symbol"] == "fblock:Calc:2"
    stress_symbols = service.list_symbols(str(FIXTURES / "stress_map.xlsx"))
    assert len(json.dumps(stress_symbols, ensure_ascii=False, indent=2)) <= RESPONSE_CHARACTER_CAP
    stress_map = service.open_workbook(str(FIXTURES / "stress_map.xlsx"))
    assert len(json.dumps(stress_map, ensure_ascii=False, indent=2)) <= RESPONSE_CHARACTER_CAP


def test_invocation_bounds_oversized_canonical_error() -> None:
    workbook = FIXTURES / "cross_sheet_model.xlsx"

    result = asyncio.run(
        server_app._invoke(
            ToolService().get_region_schema,
            str(workbook),
            "region:" + "x" * 12_000,
        )
    )

    assert result["error"]["code"] == ErrorCode.UNKNOWN_SYMBOL.value
    assert result["truncated"] is True
    assert len(json.dumps(result, ensure_ascii=False, indent=2)) <= RESPONSE_CHARACTER_CAP


def test_cell_symbol_query_finds_late_match_and_reports_complete_total(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data"
    for row in range(1, 1_101):
        sheet.cell(row, 1, row)
    path = tmp_path / "late-symbol.xlsx"
    workbook.save(path)
    service = ToolService()

    late = service.list_symbols(str(path), query="A1100", kinds=["cells"])
    all_cells = service.list_symbols(str(path), kinds=["cells"])

    assert [item["id"] for item in late["symbols"]] == ["cell:Data!A1100"]
    assert late["total"] == 1
    assert late["truncated"] is False
    assert all_cells["total"] == 1_100
    assert all_cells["truncated"] is True
    assert len(all_cells["symbols"]) <= 100


def test_value_find_reports_stale_cached_match_after_write(tmp_path: Path) -> None:
    workbook = _copy_fixture(tmp_path, "cross_sheet_model.xlsx")
    service = ToolService()

    service.write_cells(
        str(workbook),
        [WriteCellInput.model_validate({"ref": "Inputs!B2", "value": 7})],
    )
    result = service.find(str(workbook), r"^600$", search_in=["values"], sheet="Calc")

    assert result["matches"][0]["ref"] == "cell:Calc!C2"
    assert result["stale"] is True


def test_sparse_column_profiles_distinguish_range_size_from_nonnull(tmp_path: Path) -> None:
    direct_book = Workbook()
    direct_sheet = direct_book.active
    assert direct_sheet is not None
    direct_sheet.title = "Direct"
    direct_sheet["A1"] = 1
    direct_sheet["A5"] = 5
    direct_path = tmp_path / "direct-sparse.xlsx"
    direct_book.save(direct_path)

    semantic_book = Workbook()
    semantic_sheet = semantic_book.active
    assert semantic_sheet is not None
    semantic_sheet.title = "Semantic"
    semantic_sheet["A1"] = "Value"
    semantic_sheet["A2"] = 2
    semantic_sheet["A4"] = 4
    semantic_path = tmp_path / "semantic-sparse.xlsx"
    semantic_book.save(semantic_path)
    service = ToolService()

    direct = service.profile_column(str(direct_path), "Direct!A1:A5")
    semantic = service.profile_column(str(semantic_path), "col:Semantic:0:value")

    assert direct["range"] == "A1:A5"
    assert direct["count"] == 5
    assert direct["nonnull"] == 2
    assert semantic["range"] == "A2:A4"
    assert semantic["count"] == 3
    assert semantic["nonnull"] == 2


def test_cold_open_reports_each_sheet_during_index_and_cap_reduction_terminates(
    tmp_path: Path,
) -> None:
    workbook = _copy_fixture(tmp_path, "cross_sheet_model.xlsx")
    events: list[tuple[int, int, str]] = []

    result = ToolService().open_workbook(
        str(workbook), progress=lambda number, total, sheet: events.append((number, total, sheet))
    )

    assert result["reindexed"] is True
    assert events == [(1, 3, "Inputs"), (2, 3, "Calc"), (3, 3, "Summary")]
    oversized = {"items": [{"text": "x" * 20_000}]}
    fitted = service_module._fit_payload(oversized, list_keys=("items",))
    assert fitted["truncated"] is True
    assert len(json.dumps(fitted, ensure_ascii=False, indent=2)) <= RESPONSE_CHARACTER_CAP


def test_regex_engine_interrupts_catastrophic_nonmatching_subject(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data"
    sheet["A1"] = "a" * 1_000
    path = tmp_path / "regex.xlsx"
    workbook.save(path)
    started = time.monotonic()

    result = ToolService().find(str(path), r"^(a|aa)+b$", search_in=["values"])

    assert time.monotonic() - started < 3.5
    assert result["truncated"] is True
    assert result["warnings"][0]["code"] == "W_REGEX_TIMEOUT"


def test_wide_region_omits_samples_before_value_cap(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Wide"
    for column in range(1, 182):
        sheet.cell(1, column, f"Header {column}")
        sheet.cell(2, column, column)
    path = tmp_path / "wide.xlsx"
    workbook.save(path)

    result = ToolService().get_region_schema(str(path), "region:Wide:0")

    assert result["samples"] == []
    assert result["samplesOmitted"] == "181 columns; use read_range"
    assert len(json.dumps(result, ensure_ascii=False, indent=2)) <= RESPONSE_CHARACTER_CAP


def test_formula_explanations_retain_names_and_structured_references() -> None:
    service = ToolService()

    named = service.explain_formula(str(FIXTURES / "named_ranges.xlsx"), "Inputs!B4")
    structured = service.explain_formula(str(FIXTURES / "structured_table.xlsx"), "Structured!D2")

    assert named["resolvedNames"] == ["name:BaseAmount", "name:GlobalRate"]
    assert structured["structuredRefs"] == [
        "structured:Table1[Price]",
        "structured:Table1[Qty]",
    ]


def test_maximum_write_executes_all_cells_while_response_remains_bounded(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data"
    sheet["A1"] = 0
    path = tmp_path / "maximum-write.xlsx"
    workbook.save(path)
    cells = [
        WriteCellInput.model_validate({"ref": f"Data!A{row}", "value": row})
        for row in range(1, 501)
    ]

    result = ToolService().write_cells(str(path), cells)

    assert result["resultsTotal"] == 500
    assert result["truncated"] is True
    assert len(result["results"]) < 500
    assert len(json.dumps(result, ensure_ascii=False, indent=2)) <= RESPONSE_CHARACTER_CAP
    reloaded = load_workbook(path, read_only=True, data_only=True)
    assert reloaded["Data"]["A500"].value == 500
    reloaded.close()


def test_root_allowlist_follows_real_paths_and_denies_escape(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    allowed = tmp_path / "allowed"
    allowed.mkdir()
    workbook = _copy_fixture(allowed, "basic_single_table.xlsx")
    outside = _copy_fixture(tmp_path, "cross_sheet_model.xlsx")
    monkeypatch.setenv("EXCEL_LSP_ROOT", str(allowed))
    service = ToolService()

    assert service.open_workbook(str(workbook))["workbook"] == workbook.name
    with pytest.raises(ExcelLSPError) as caught:
        service.open_workbook(str(outside))
    assert caught.value.code is ErrorCode.PATH_DENIED

    link = allowed / "escape.xlsx"
    try:
        link.symlink_to(outside)
    except OSError:
        pytest.skip("symlink creation is not available")
    with pytest.raises(ExcelLSPError) as caught_link:
        service.open_workbook(str(link))
    assert caught_link.value.code is ErrorCode.PATH_DENIED


@pytest.mark.parametrize(
    "payload",
    (
        {"ref": "Sales!A1"},
        {"ref": "Sales!A1", "value": 1, "formula": "=1"},
        {"ref": "Sales!A1", "formula": "1+1"},
    ),
)
def test_write_tool_shapes_invalid_operation_as_per_cell_error(
    tmp_path: Path, payload: dict[str, object]
) -> None:
    workbook = _copy_fixture(tmp_path, "basic_single_table.xlsx")
    cell = WriteCellInput.model_validate(payload)

    result = ToolService().write_cells(str(workbook), [cell])

    assert result["results"][0]["ok"] is False
    assert result["results"][0]["error"]["code"] == ErrorCode.INVALID_VALUE.value
